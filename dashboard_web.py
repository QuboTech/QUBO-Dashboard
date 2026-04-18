"""
dashboard_web.py - QUBO Dashboard v4
=====================================
Dashboard principal focado nos produtos da loja.
P�gina /config separada para uploads, ML Auth e configurações.

Autor: Claude para QUBO
Data: 2026-04
"""
import os, json, io, time, threading, logging
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, send_file, session, redirect
from db import get_conn, garantir_schema, USAR_POSTGRES
from auth import (login_required, admin_required, get_tenant_id, get_usuario_nome,
                  get_tenant_nome, verificar_login, criar_tenant_e_admin,
                  convidar_usuario, listar_usuarios_tenant, remover_usuario_tenant,
                  get_tenant_info, LOGIN_HTML, SIGNUP_HTML)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "vitrix-secret-2026")
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Filtro Jinja ─────────────────────────────────────────────────────
@app.template_filter('br')
def fmt_br(v, d=2):
    try:
        return f"{float(v):,.{d}f}".replace(",","X").replace(".",",").replace("X",".")
    except:
        return "0,00"

# ── Tabela de frete ML (110 faixas preço x peso) ─────────────────────
FRETE = [
    (79,99.99,0,0.299,11.97),(79,99.99,0.3,0.499,12.87),(79,99.99,0.5,0.999,13.47),
    (79,99.99,1,1.999,14.07),(79,99.99,2,2.999,14.97),(79,99.99,3,3.999,16.17),
    (79,99.99,4,4.999,17.07),(79,99.99,5,8.999,26.67),
    (100,119.99,0,0.299,13.97),(100,119.99,0.3,0.499,15.02),(100,119.99,0.5,0.999,15.72),
    (100,119.99,1,1.999,16.42),(100,119.99,2,2.999,17.47),(100,119.99,3,3.999,18.87),
    (100,119.99,4,4.999,19.92),(100,119.99,5,8.999,31.12),
    (120,149.99,0,0.299,15.96),(120,149.99,0.3,0.499,17.16),(120,149.99,0.5,0.999,17.96),
    (120,149.99,1,1.999,18.76),(120,149.99,2,2.999,19.96),(120,149.99,3,3.999,21.56),
    (120,149.99,4,4.999,22.76),(120,149.99,5,8.999,35.56),
    (150,199.99,0,0.299,17.96),(150,199.99,0.3,0.499,19.31),(150,199.99,0.5,0.999,20.21),
    (150,199.99,1,1.999,21.11),(150,199.99,2,2.999,22.46),(150,199.99,3,3.999,24.26),
    (150,199.99,4,4.999,25.61),(150,199.99,5,8.999,40.01),
    (200,9999,0,0.299,19.95),(200,9999,0.3,0.499,21.45),(200,9999,0.5,0.999,22.45),
    (200,9999,1,1.999,23.45),(200,9999,2,2.999,24.95),(200,9999,3,3.999,26.95),
    (200,9999,4,4.999,28.45),(200,9999,5,8.999,44.45),
]

def calcular_frete(preco, peso):
    if not preco or preco < 79.90 or not peso: return 0.0
    for pm,px,wm,wx,fr in FRETE:
        if pm <= preco <= px and wm <= peso <= wx: return fr
    return 0.0

def calcular(custo, preco_ml, taxa_pct=16.5, peso=0, emb=0, imp_pct=0):
    if not custo or not preco_ml or preco_ml <= 0:
        return dict(custo_total=0,margem_pct=0,margem_r=0,viavel=0,frete=0,taxa_fixa=0)
    taxa = taxa_pct/100; imp = imp_pct/100
    taxa_fixa = 6.25 if preco_ml < 79 else 0
    frete = calcular_frete(preco_ml, peso)
    custo_total = custo + preco_ml*taxa + taxa_fixa + preco_ml*imp + emb + frete
    margem_r = round(preco_ml - custo_total, 2)
    margem_pct = round(margem_r / preco_ml * 100, 1) if preco_ml > 0 else 0
    return dict(custo_total=round(custo_total,2), margem_pct=margem_pct,
                margem_r=margem_r, viavel=1 if margem_pct >= 20 else 0,
                frete=round(frete,2), taxa_fixa=taxa_fixa)

def get_aliquota():
    try:
        conn = get_conn(); cur = conn.cursor()
        ph = "%s" if USAR_POSTGRES else "?"
        cur.execute(f"SELECT valor FROM config WHERE chave = {ph}", ("aliquota_imposto",))
        r = cur.fetchone(); conn.close()
        return float(r[0]) if r else 0.0
    except: return 0.0

def ph(): return "%s" if USAR_POSTGRES else "?"

def _brand_ctx():
    """Contexto de branding pros templates. Qubo tenant vê 'Qubo'; demais veem 'Vitrix'."""
    tid = get_tenant_id()
    return {
        'brand': 'Qubo' if tid == 'qubo' else 'Vitrix',
        'plataforma': 'Vitrix',
        'tenant_nome': get_tenant_nome(),
        'tenant_slug': tid,
        'is_admin': session.get('role') == 'admin',
    }

# ════════════════════════════════════════════════════════════════════
# AUTH
# ════════════════════════════════════════════════════════════════════
@app.route('/login')
def pagina_login():
    return LOGIN_HTML

@app.route('/signup')
def pagina_signup():
    return SIGNUP_HTML

@app.route('/api/login', methods=['POST'])
def api_login():
    d = request.get_json() or {}
    u = d.get('usuario','').strip()
    s = d.get('senha','')
    user = verificar_login(u, s)
    if user:
        session['usuario'] = u
        session['usuario_nome'] = user.get('nome', u)
        session['tenant_id'] = user.get('tenant_id', 'qubo')
        session['tenant_slug'] = user.get('tenant_slug', user.get('tenant_id', 'qubo'))
        session['tenant_nome'] = user.get('tenant_nome', 'Vitrix')
        session['role'] = user.get('role', 'admin')
        session['fonte_auth'] = user.get('fonte', 'env')
        session.permanent = True
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'erro': 'Email/usuário ou senha incorretos'})

@app.route('/api/signup', methods=['POST'])
def api_signup():
    d = request.get_json() or {}
    res = criar_tenant_e_admin(
        nome_empresa=d.get('nome_empresa', ''),
        email=d.get('email', ''),
        senha=d.get('senha', ''),
        nome_admin=d.get('nome', '')
    )
    if not res.get('ok'):
        return jsonify(res)
    # Auto-login após signup
    session['usuario'] = d.get('email', '').strip()
    session['usuario_nome'] = d.get('nome', '') or d.get('email', '').split('@')[0]
    session['tenant_id'] = res['tenant_slug']
    session['tenant_slug'] = res['tenant_slug']
    session['tenant_nome'] = res['nome_empresa']
    session['role'] = 'admin'
    session['fonte_auth'] = 'db'
    session.permanent = True
    return jsonify({'ok': True, 'tenant_slug': res['tenant_slug']})

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# ────── Admin de usuários do tenant ──────
@app.route('/api/admin/usuarios')
@admin_required
def api_admin_usuarios():
    tid = get_tenant_id()
    return jsonify({'ok': True, 'usuarios': listar_usuarios_tenant(tid)})

@app.route('/api/admin/convidar', methods=['POST'])
@admin_required
def api_admin_convidar():
    d = request.get_json() or {}
    tid = get_tenant_id()
    return jsonify(convidar_usuario(
        tenant_slug=tid,
        email=d.get('email',''),
        senha=d.get('senha',''),
        nome=d.get('nome',''),
        role=d.get('role','user')
    ))

@app.route('/api/admin/remover', methods=['POST'])
@admin_required
def api_admin_remover():
    d = request.get_json() or {}
    tid = get_tenant_id()
    return jsonify(remover_usuario_tenant(tid, int(d.get('usuario_id', 0))))

@app.route('/api/tenant-info')
@login_required
def api_tenant_info():
    tid = get_tenant_id()
    info = get_tenant_info(tid)
    info['usuario_nome'] = get_usuario_nome()
    info['role'] = session.get('role', 'admin')
    return jsonify({'ok': True, **info})

# ════════════════════════════════════════════════════════════════════
# DASHBOARD PRINCIPAL
# ════════════════════════════════════════════════════════════════════
@app.route('/')
@login_required
def index():
    conn = get_conn(); cur = conn.cursor()
    tenant = get_tenant_id()
    p = ph()

    # Filtros
    f = {
        'forn': request.args.get('forn',''),
        'produto': request.args.get('produto',''),
        'status': request.args.get('status',''),
        'cmin': request.args.get('cmin','0'),
        'cmax': request.args.get('cmax','999'),
        'pp': int(request.args.get('pp', 100)),
    }
    pg = max(1, int(request.args.get('pg', 1)))

    w = [f"tenant_id = {p}"]; params = [tenant]
    if f['forn']: w.append(f"fornecedor ILIKE {p}" if USAR_POSTGRES else f"fornecedor LIKE {p}"); params.append(f"%{f['forn']}%")
    if f['produto']: w.append(f"descricao ILIKE {p}" if USAR_POSTGRES else f"descricao LIKE {p}"); params.append(f"%{f['produto']}%")
    if f['status'] == 'viavel': w.append("viavel = 1")
    elif f['status'] == 'nao_viavel': w.append("viavel = 0 AND preco_ml > 0")
    elif f['status'] == 'pendente': w.append("preco_ml = 0 OR preco_ml IS NULL")
    elif f['status'] == 'escolhido': w.append("escolhido = 1")
    try:
        w.append(f"custo >= {p}"); params.append(float(f['cmin']))
        w.append(f"custo <= {p}"); params.append(float(f['cmax']))
    except: pass

    where = " AND ".join(w)
    cur.execute(f"SELECT COUNT(*) FROM produtos WHERE {where}", params)
    total_f = cur.fetchone()[0]
    total_pg = max(1, (total_f + f['pp'] - 1) // f['pp'])
    pg = min(pg, total_pg)

    cur.execute(f"SELECT * FROM produtos WHERE {where} ORDER BY id DESC LIMIT {p} OFFSET {p}",
                params + [f['pp'], (pg-1)*f['pp']])
    cols = [d[0] for d in cur.description]
    produtos = [dict(zip(cols, r)) for r in cur.fetchall()]

    # Stats
    cur.execute(f"SELECT COUNT(*) FROM produtos WHERE tenant_id = {p}", [tenant]); total = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(DISTINCT fornecedor) FROM produtos WHERE tenant_id = {p}", [tenant]); forn_c = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM produtos WHERE tenant_id = {p} AND preco_ml > 0", [tenant]); c_preco = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM produtos WHERE tenant_id = {p} AND viavel = 1", [tenant]); c_viavel = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM produtos WHERE tenant_id = {p} AND viavel = 0 AND preco_ml > 0", [tenant]); c_nviavel = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM produtos WHERE tenant_id = {p} AND (preco_ml = 0 OR preco_ml IS NULL)", [tenant]); c_pend = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM produtos WHERE tenant_id = {p} AND escolhido = 1", [tenant]); c_escolhido = cur.fetchone()[0]
    cur.execute(f"SELECT DISTINCT fornecedor FROM produtos WHERE tenant_id = {p} ORDER BY fornecedor", [tenant])
    fornecedores = [r[0] for r in cur.fetchall() if r[0]]
    conn.close()

    stats = dict(total=total, fornecedores=forn_c, com_preco=c_preco, sem_preco=total-c_preco,
                 viaveis=c_viavel, nao_viaveis=c_nviavel, pendentes=c_pend, escolhidos=c_escolhido)

    imp_pct = round(get_aliquota()*100, 1)
    usuario = get_usuario_nome()

    def url_pg(p_num):
        args = request.args.to_dict(); args['pg'] = p_num
        return '?' + '&'.join(f"{k}={v}" for k,v in args.items())

    return render_template_string(HTML_DASH, produtos=produtos, stats=stats,
        fornecedores=fornecedores, f=f, pg=pg, total_paginas=total_pg,
        url_pg=url_pg, imp_pct=imp_pct, usuario=usuario, total_filtrado=total_f,
        **_brand_ctx())

# ════════════════════════════════════════════════════════════════════
# APIS CRUD PRODUTOS
# ════════════════════════════════════════════════════════════════════
@app.route('/api/atualizar', methods=['POST'])
@login_required
def api_atualizar():
    d = request.get_json(); pid = d.get('id'); campo = d.get('campo'); valor = d.get('valor')
    if not all([pid, campo]): return jsonify({'ok': False})
    campos_ok = ['preco_ml','taxa_categoria','peso_kg','custo_embalagem','custo_ads',
                 'custo_full','link_ml','notas','tipo_anuncio','promo_percentual','margem_alvo']
    if campo not in campos_ok: return jsonify({'ok': False, 'erro': 'Campo não permitido'})
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph(); tenant = get_tenant_id()
        cur.execute(f"UPDATE produtos SET {campo} = {p} WHERE id = {p} AND tenant_id = {p}", [valor, pid, tenant])
        # Recalcula se tem preco_ml
        cur.execute(f"SELECT custo, preco_ml, taxa_categoria, peso_kg, custo_embalagem FROM produtos WHERE id = {p}", [pid])
        row = cur.fetchone()
        if row:
            custo, preco_ml, taxa, peso, emb = row
            imp_pct = get_aliquota() * 100
            r = calcular(custo or 0, preco_ml or 0, (taxa or 0.165)*100, peso or 0, emb or 0, imp_pct)
            cur.execute(f"""UPDATE produtos SET custo_total={p}, margem_percentual={p},
                margem_reais={p}, viavel={p} WHERE id = {p}""",
                [r['custo_total'], r['margem_pct'], r['margem_r'], r['viavel'], pid])
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/adicionar-produto', methods=['POST'])
@login_required
def api_adicionar_produto():
    d = request.get_json(); tenant = get_tenant_id()
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        cur.execute(f"""INSERT INTO produtos (tenant_id, fornecedor, descricao, custo, codigo)
            VALUES ({p},{p},{p},{p},{p}) RETURNING id""" if USAR_POSTGRES else
            f"""INSERT INTO produtos (tenant_id, fornecedor, descricao, custo, codigo)
            VALUES ({p},{p},{p},{p},{p})""",
            [tenant, d.get('fornecedor',''), d.get('descricao',''), float(d.get('custo',0)), d.get('codigo','')])
        if USAR_POSTGRES: pid = cur.fetchone()[0]
        else: pid = cur.lastrowid
        conn.commit(); conn.close()
        return jsonify({'ok': True, 'id': pid, 'desc': d.get('descricao','')[:25]})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/deletar-produto', methods=['POST'])
@login_required
def api_deletar_produto():
    d = request.get_json(); pid = d.get('id'); tenant = get_tenant_id()
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        cur.execute(f"DELETE FROM produtos WHERE id = {p} AND tenant_id = {p}", [pid, tenant])
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/escolher', methods=['POST'])
@login_required
def api_escolher():
    d = request.get_json(); tenant = get_tenant_id()
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        cur.execute(f"UPDATE produtos SET escolhido = 1 WHERE id = {p} AND tenant_id = {p}", [d.get('id'), tenant])
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/descartar', methods=['POST'])
@login_required
def api_descartar():
    d = request.get_json(); tenant = get_tenant_id()
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        cur.execute(f"UPDATE produtos SET escolhido = 0 WHERE id = {p} AND tenant_id = {p}", [d.get('id'), tenant])
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/aliquota', methods=['POST'])
@login_required
def api_aliquota():
    d = request.get_json()
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        val = float(d.get('valor', 0)) / 100
        if USAR_POSTGRES:
            cur.execute(f"INSERT INTO config (chave, valor) VALUES ({p},{p}) ON CONFLICT (chave) DO UPDATE SET valor={p}",
                       ['aliquota_imposto', str(val), str(val)])
        else:
            cur.execute(f"INSERT OR REPLACE INTO config (chave, valor) VALUES ({p},{p})", ['aliquota_imposto', str(val)])
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/recalcular-todos', methods=['POST'])
@login_required
def api_recalcular_todos():
    tenant = get_tenant_id(); imp_pct = get_aliquota() * 100
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        cur.execute(f"SELECT id, custo, preco_ml, taxa_categoria, peso_kg, custo_embalagem FROM produtos WHERE tenant_id = {p}", [tenant])
        rows = cur.fetchall(); atualizados = 0
        for row in rows:
            pid, custo, preco_ml, taxa, peso, emb = row
            if not custo: continue
            r = calcular(custo or 0, preco_ml or 0, (taxa or 0.165)*100, peso or 0, emb or 0, imp_pct)
            cur.execute(f"UPDATE produtos SET custo_total={p}, margem_percentual={p}, margem_reais={p}, viavel={p} WHERE id={p}",
                       [r['custo_total'], r['margem_pct'], r['margem_r'], r['viavel'], pid])
            atualizados += 1
        conn.commit(); conn.close()
        return jsonify({'ok': True, 'atualizados': atualizados})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/buscar-ml', methods=['POST'])
@login_required
def api_buscar_ml():
    """Busca preço médio no ML. A busca pública não exige token — usa token apenas para taxa real."""
    import requests as req
    from agente_pesquisa import limpar_termo_busca
    d = request.get_json(); termo = d.get('termo','')
    if not termo: return jsonify({'ok': False, 'erro': 'Termo vazio'})
    try:
        # Token é opcional — search ML é API pública
        headers = {}
        try:
            from ml_buscador import MLBuscador
            ml = MLBuscador(tenant_id=get_tenant_id())
            if ml.esta_autenticado():
                headers = {"Authorization": f"Bearer {ml.auth.access_token}"}
        except Exception:
            pass

        # Limpa o termo antes de buscar
        termo_limpo = limpar_termo_busca(termo) or termo[:50]

        r = req.get("https://api.mercadolibre.com/sites/MLB/search",
                   headers=headers, params={"q": termo_limpo, "limit": 10}, timeout=15)

        # Se 401/403 com token, tenta sem token
        if r.status_code in (401, 403) and headers:
            headers = {}
            r = req.get("https://api.mercadolibre.com/sites/MLB/search",
                       params={"q": termo_limpo, "limit": 10}, timeout=15)

        if r.status_code != 200: return jsonify({'ok': False, 'erro': f'Erro ML: {r.status_code}'})
        results = r.json().get('results', [])
        if not results: return jsonify({'ok': False, 'erro': f'Sem resultados para "{termo_limpo}"'})
        precos = [x.get('price',0) for x in results if x.get('price',0) > 0]
        preco_medio = round(sum(precos)/len(precos), 2) if precos else 0
        # Taxa real (só com token; sem token usa default)
        taxa_pct = 16.5
        cat = results[0].get('category_id','')
        if cat and headers:
            try:
                r2 = req.get("https://api.mercadolibre.com/sites/MLB/listing_prices",
                            headers=headers, params={"price": preco_medio, "category_id": cat, "currency_id": "BRL"}, timeout=10)
                if r2.status_code == 200:
                    for lst in r2.json():
                        if lst.get('listing_type_id') in ('gold_special','gold_pro'):
                            for comp in lst.get('sale_fee_components',[]):
                                if comp.get('type') == 'fee': taxa_pct = round(comp.get('ratio',0.165)*100, 1)
            except Exception:
                pass
        return jsonify({'ok': True, 'preco_medio': preco_medio, 'taxa_percentual': taxa_pct,
                       'total': len(results), 'termo_buscado': termo_limpo})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/atualizar-taxas-ml', methods=['POST'])
@login_required
def api_atualizar_taxas_ml():
    return jsonify({'ok': True, 'atualizados': 0})

@app.route('/exportar')
@login_required
def exportar():
    import openpyxl; from openpyxl.styles import PatternFill, Font, Alignment
    tenant = get_tenant_id()
    conn = get_conn(); cur = conn.cursor(); p = ph()
    cur.execute(f"SELECT * FROM produtos WHERE tenant_id = {p} ORDER BY fornecedor, descricao", [tenant])
    cols = [d[0] for d in cur.description]; rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Produtos QUBO"
    headers = ['ID','Fornecedor','Código','Descrição','Custo','Preço ML','Taxa %','Frete','Custo Total','Margem %','Margem R$','Viável','Escolhido']
    col_map = {'id':0,'fornecedor':1,'codigo':2,'descricao':3,'custo':4,'preco_ml':5,'taxa_categoria':6,'custo_frete':7,'custo_total':8,'margem_percentual':9,'margem_reais':10,'viavel':11,'escolhido':12}
    header_fill = PatternFill(fgColor="1a1f3a", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h); cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal='center')
    for row in rows:
        d = dict(zip(cols, row))
        ws.append([d.get('id'), d.get('fornecedor'), d.get('codigo'), d.get('descricao'),
                   d.get('custo'), d.get('preco_ml'), round((d.get('taxa_categoria') or 0)*100, 1),
                   d.get('custo_frete'), d.get('custo_total'), d.get('margem_percentual'),
                   d.get('margem_reais'), 'SIM' if d.get('viavel') else 'NÃO',
                   'SIM' if d.get('escolhido') else ''])
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'qubo_produtos_{datetime.now().strftime("%Y%m%d")}.xlsx')

# ════════════════════════════════════════════════════════════════════
# AGENTES IA
# ════════════════════════════════════════════════════════════════════
@app.route('/api/pesquisar-produto', methods=['POST'])
@login_required
def api_pesquisar_produto():
    try:
        from agente_pesquisa import analisar_produto_ml
        from ml_buscador import MLBuscador
        d = request.get_json()
        token = None
        try:  # token é opcional — busca pública funciona sem ele
            ml = MLBuscador(tenant_id=get_tenant_id())
            if ml.esta_autenticado():
                token = ml.auth.access_token
        except Exception:
            pass
        return jsonify(analisar_produto_ml(d.get('id'), token))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

def _get_ml_token():
    """Retorna (token, user_id) do ML do TENANT ATUAL. Nunca lança."""
    try:
        from ml_buscador import MLBuscador
        tid = get_tenant_id()
        ml = MLBuscador(tenant_id=tid)
        if ml.esta_autenticado():
            return ml.auth.access_token, ml.auth.user_id
    except Exception:
        pass
    return None, None


@app.route('/api/sugerir-taxa', methods=['POST'])
@login_required
def api_sugerir_taxa():
    """Sugere taxa ML correta por categoria/descrição. Token melhora precisão mas é opcional."""
    try:
        from taxas_ml import sugerir_taxa_produto
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        return jsonify(sugerir_taxa_produto(
            descricao=d.get('descricao', ''),
            preco=float(d.get('preco', 0)),
            category_id=d.get('category_id', ''),
            token_ml=token or ''
        ))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})


@app.route('/api/precificar', methods=['POST'])
@login_required
def api_precificar():
    try:
        from agente_precificacao import precificar_produto
        d = request.get_json()
        token, _ = _get_ml_token()  # token opcional — melhora taxa se disponível
        return jsonify(precificar_produto(d.get('id'), token or "",
                       float(d.get('margem_minima',20)), float(d.get('imposto_pct',0))))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/viabilidade', methods=['POST'])
@login_required
def api_viabilidade():
    try:
        from agente_viabilidade import analisar_viabilidade
        d = request.get_json()
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        return jsonify(analisar_viabilidade(d.get('termo',''), token,
            float(d.get('custo',0)), float(d.get('peso',0)), float(d.get('embalagem',0)),
            float(d.get('imposto',0)), float(d.get('margem_minima',20)), d.get('id')))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/alerta-diario')
@login_required
def api_alerta_diario():
    try:
        from agente_alerta import gerar_alerta_diario
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        return jsonify(gerar_alerta_diario(token, user_id))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/tendencias')
@login_required
def api_tendencias():
    try:
        from agente_tendencias import buscar_tendencias
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        return jsonify(buscar_tendencias(token))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/saude-anuncios')
@login_required
def api_saude_anuncios():
    try:
        from agente_saude import monitorar_saude
        from ml_buscador import MLBuscador
        ml = MLBuscador(tenant_id=get_tenant_id())
        if not ml.esta_autenticado(): return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(monitorar_saude(ml.auth.access_token, ml.auth.user_id))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/pedidos')
@login_required
def api_pedidos():
    try:
        from agente_pedidos import listar_pedidos
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        dias = int(request.args.get('dias', 30))
        status = request.args.get('status', '')
        return jsonify(listar_pedidos(token, user_id, dias, status))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/perguntas')
@login_required
def api_perguntas():
    try:
        from agente_perguntas import listar_perguntas
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        apenas = request.args.get('todas', '0') != '1'
        return jsonify(listar_perguntas(token, user_id, apenas_nao_respondidas=apenas))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/responder-pergunta', methods=['POST'])
@login_required
def api_responder_pergunta():
    try:
        from agente_perguntas import responder_pergunta
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(responder_pergunta(token, d.get('question_id'), d.get('resposta', '')))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/meus-anuncios')
@login_required
def api_meus_anuncios():
    try:
        from agente_anuncios import listar_anuncios
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        status = request.args.get('status', 'active')
        offset = int(request.args.get('offset', 0))
        return jsonify(listar_anuncios(token, user_id, status, 50, offset))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/pausar-anuncio', methods=['POST'])
@login_required
def api_pausar_anuncio():
    try:
        from agente_anuncios import pausar_anuncio
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(pausar_anuncio(token, d.get('item_id', '')))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/ativar-anuncio', methods=['POST'])
@login_required
def api_ativar_anuncio():
    try:
        from agente_anuncios import ativar_anuncio
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(ativar_anuncio(token, d.get('item_id', '')))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/atualizar-estoque', methods=['POST'])
@login_required
def api_atualizar_estoque():
    try:
        from agente_anuncios import atualizar_estoque
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(atualizar_estoque(token, d.get('item_id', ''), int(d.get('quantidade', 0))))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/reputacao')
@login_required
def api_reputacao():
    try:
        from agente_reputacao import obter_reputacao
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        return jsonify(obter_reputacao(token, user_id))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/faturamento')
@login_required
def api_faturamento():
    try:
        from agente_faturamento import obter_faturamento
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        periodo = request.args.get('periodo', '')
        return jsonify(obter_faturamento(token, user_id, periodo))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

# ════════════════════════════════════════════════════════════════════
# CRIAR / EDITAR ANÚNCIOS
# ════════════════════════════════════════════════════════════════════
@app.route('/api/prever-categoria', methods=['POST'])
@login_required
def api_prever_categoria():
    try:
        from agente_criar_anuncio import prever_categoria
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        return jsonify(prever_categoria(token or '', d.get('titulo', '')))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/atributos-categoria', methods=['POST'])
@login_required
def api_atributos_categoria():
    try:
        from agente_criar_anuncio import obter_atributos_categoria
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        return jsonify(obter_atributos_categoria(token or '', d.get('category_id', '')))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/validar-anuncio', methods=['POST'])
@login_required
def api_validar_anuncio():
    try:
        from agente_criar_anuncio import validar_anuncio
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(validar_anuncio(token, d))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/criar-anuncio', methods=['POST'])
@login_required
def api_criar_anuncio():
    try:
        from agente_criar_anuncio import criar_anuncio
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(criar_anuncio(token, d))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/editar-anuncio', methods=['POST'])
@login_required
def api_editar_anuncio():
    try:
        from agente_criar_anuncio import editar_anuncio
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(editar_anuncio(token, d.get('item_id', ''), d.get('alteracoes', {})))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/publicar-produto', methods=['POST'])
@login_required
def api_publicar_produto():
    """Atalho: cria anúncio a partir de produto do catálogo QUBO."""
    try:
        from agente_criar_anuncio import criar_a_partir_de_produto
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(criar_a_partir_de_produto(
            token, int(d.get('id', 0)),
            preco=float(d.get('preco', 0)) or None,
            quantidade=int(d.get('quantidade', 1)),
            imagens=d.get('imagens', []),
            tipo=d.get('tipo', 'gold_special')
        ))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

# ════════════════════════════════════════════════════════════════════
# MÉTRICAS (Visitas e Conversão)
# ════════════════════════════════════════════════════════════════════
@app.route('/api/metricas')
@login_required
def api_metricas():
    try:
        from agente_metricas import analise_completa
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        dias = int(request.args.get('dias', 30))
        limite = int(request.args.get('limite', 30))
        return jsonify(analise_completa(token, user_id, dias, limite))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

# ════════════════════════════════════════════════════════════════════
# PRICE TO WIN / BENCHMARKS
# ════════════════════════════════════════════════════════════════════
@app.route('/api/price-to-win', methods=['POST'])
@login_required
def api_price_to_win():
    try:
        from agente_price_win import price_to_win
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(price_to_win(token, d.get('item_id', '')))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/benchmark')
@login_required
def api_benchmark():
    try:
        from agente_price_win import benchmark_vendedor
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado. Conecte em Config → ML Auth'})
        limite = int(request.args.get('limite', 30))
        return jsonify(benchmark_vendedor(token, user_id, limite))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/descontos-sugeridos')
@login_required
def api_descontos_sugeridos():
    try:
        from agente_price_win import descontos_sugeridos
        token, user_id = _get_ml_token()
        if not token: return jsonify({'ok': False, 'erro': 'ML não conectado'})
        return jsonify(descontos_sugeridos(token, user_id))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

# ════════════════════════════════════════════════════════════════════
# WEBHOOKS — receptor de notificações ML em tempo real
# ════════════════════════════════════════════════════════════════════
@app.route('/webhook/ml', methods=['POST'])
def webhook_ml():
    """Endpoint público — ML envia notificações aqui. NÃO exige login."""
    try:
        from webhook_handler import processar_notificacao
        body = request.get_json(silent=True) or {}
        result = processar_notificacao(body)
        # ML exige resposta rápida (<500ms). Retorna 200 sempre.
        return jsonify({'received': True}), 200
    except Exception as e:
        logger.error(f"Webhook error: {e}")
        return jsonify({'received': True}), 200  # sempre 200 p/ ML não retry

@app.route('/api/webhooks-eventos')
@login_required
def api_webhooks_eventos():
    try:
        from webhook_handler import listar_eventos
        topic = request.args.get('topic', '')
        return jsonify(listar_eventos(limite=80, topic=topic, tenant_id=get_tenant_id()))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})


# ════════════════════════════════════════════════════════════════════
# ESPIÃO — vendas/snapshot de qualquer anúncio ML (todos os sellers)
# ════════════════════════════════════════════════════════════════════
@app.route('/api/spy-anuncio', methods=['POST'])
@login_required
def api_spy_anuncio():
    try:
        from agente_espiao import spy_anuncio
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        return jsonify(spy_anuncio(d.get('item_id', ''), token or '', tenant_id=get_tenant_id()))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/watch-add', methods=['POST'])
@login_required
def api_watch_add():
    try:
        from agente_espiao import adicionar_watch
        d = request.get_json() or {}
        token, _ = _get_ml_token()
        return jsonify(adicionar_watch(d.get('item_id', ''), d.get('apelido', ''), token or '', tenant_id=get_tenant_id()))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/watch-remove', methods=['POST'])
@login_required
def api_watch_remove():
    try:
        from agente_espiao import remover_watch
        d = request.get_json() or {}
        return jsonify(remover_watch(d.get('item_id', ''), tenant_id=get_tenant_id()))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/watchlist')
@login_required
def api_watchlist():
    try:
        from agente_espiao import listar_watchlist
        token, _ = _get_ml_token()
        refresh = request.args.get('refresh', '1') == '1'
        return jsonify(listar_watchlist(token or '', refresh=refresh, tenant_id=get_tenant_id()))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/ranking-busca')
@login_required
def api_ranking_busca():
    try:
        from agente_espiao import ranking_busca
        query = request.args.get('q', '').strip()
        limite = int(request.args.get('limite', 30))
        token, _ = _get_ml_token()
        return jsonify(ranking_busca(query, limite, token or ''))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/ranking-categoria')
@login_required
def api_ranking_categoria():
    try:
        from agente_espiao import ranking_categoria
        cat = request.args.get('cat', '').strip()
        limite = int(request.args.get('limite', 30))
        token, _ = _get_ml_token()
        return jsonify(ranking_categoria(cat, limite, token or ''))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/spy-snapshot-cron', methods=['POST'])
@login_required
def api_spy_snapshot_cron():
    try:
        from agente_espiao import snapshot_cron
        token, _ = _get_ml_token()
        return jsonify(snapshot_cron(token or '', tenant_id=get_tenant_id()))
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})


# ════════════════════════════════════════════════════════════════════
# PÁGINA DE CONFIGURAÇÕES
# ════════════════════════════════════════════════════════════════════
@app.route('/ml')
@login_required
def painel_ml():
    usuario = session.get('usuario', '')
    return render_template_string(HTML_ML, usuario=usuario, **_brand_ctx())


@app.route('/config')
@login_required
def pagina_config():
    from ml_buscador import MLBuscador
    try:
        ml = MLBuscador(tenant_id=get_tenant_id()); ml_ok = ml.esta_autenticado(); ml_user = ml.auth.user_id if ml_ok else None
    except: ml_ok = False; ml_user = None
    usuario = get_usuario_nome()
    return render_template_string(HTML_CONFIG, ml_ok=ml_ok, ml_user=ml_user, usuario=usuario, **_brand_ctx())

@app.route('/admin/usuarios')
@admin_required
def pagina_admin_usuarios():
    usuario = get_usuario_nome()
    tenant = get_tenant_info(get_tenant_id())
    return render_template_string(HTML_ADMIN_USUARIOS, usuario=usuario, tenant=tenant, **_brand_ctx())

@app.route('/api/ml-auth', methods=['POST'])
@login_required
def api_ml_auth():
    from ml_buscador import MLBuscador
    d = request.get_json(); code = d.get('code','').strip()
    if not code: return jsonify({'ok': False, 'erro': 'Código obrigatório'})
    try:
        ml = MLBuscador(tenant_id=get_tenant_id())
        result = ml.trocar_codigo(code)
        return jsonify({'ok': result.get('ok', False),
                       'user_id': result.get('user_id'),
                       'erro': result.get('erro')})
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/ml-status')
@login_required
def api_ml_status():
    try:
        from ml_buscador import MLBuscador
        ml = MLBuscador(tenant_id=get_tenant_id())
        auth = ml.auth
        exp_str = auth.expires_at.isoformat() if auth.expires_at else None
        return jsonify({
            'ok': True,
            'autenticado': ml.esta_autenticado(),
            'user_id': auth.user_id,
            'tem_access_token': bool(auth.access_token),
            'tem_refresh_token': bool(auth.refresh_token),
            'expires_at': exp_str,
        })
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})


@app.route('/api/ml-debug')
@login_required
def api_ml_debug():
    """Diagnóstico do estado do token ML no banco e em memória."""
    try:
        from db import get_conn, USAR_POSTGRES
        conn = get_conn(); cur = conn.cursor()
        ph = "%s" if USAR_POSTGRES else "?"
        cur.execute(f"SELECT id, user_id, expires_at, salvo_em, length(access_token), length(refresh_token) FROM ml_tokens WHERE id = {ph}", (get_tenant_id(),))
        row = cur.fetchone(); conn.close()
        db_info = None
        if row:
            db_info = {'user_id': row[2], 'expires_at': row[3], 'salvo_em': row[4], 'len_access': row[5], 'len_refresh': row[6]}
        from ml_buscador import MLBuscador, _carregar_token_db
        loaded = _carregar_token_db(get_tenant_id())
        ml = MLBuscador(tenant_id=get_tenant_id())
        from datetime import datetime
        return jsonify({
            'ok': True,
            'db': db_info,
            'carregado': {
                'tem_access': bool(loaded.get('access_token')),
                'tem_refresh': bool(loaded.get('refresh_token')),
                'expires_at': loaded.get('expires_at'),
            },
            'memoria': {
                'autenticado': ml.esta_autenticado(),
                'tem_access': bool(ml.auth.access_token),
                'tem_refresh': bool(ml.auth.refresh_token),
                'expires_at': ml.auth.expires_at.isoformat() if ml.auth.expires_at else None,
                'agora_servidor': datetime.now().isoformat(),
            }
        })
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'erro': str(e), 'trace': traceback.format_exc()})

@app.route('/api/watcher-upload', methods=['POST'])
def api_watcher_upload():
    """Endpoint para o File Watcher local — autenticado por API key, sem sessão."""
    import tempfile, os as _os
    api_key = request.headers.get('X-API-Key', '')
    expected = os.getenv('WATCHER_API_KEY', 'qubo-watcher-2026')
    if api_key != expected:
        return jsonify({'ok': False, 'erro': 'Chave de API inválida'}), 401
    if 'arquivo' not in request.files:
        return jsonify({'ok': False, 'erro': 'Nenhum arquivo enviado'})
    arquivo = request.files['arquivo']
    if not arquivo.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'erro': 'Apenas arquivos PDF'})
    tenant = request.form.get('tenant_id', 'gustavo')
    try:
        from multi_extractor import MultiExtractor
        fornecedor = _os.path.splitext(arquivo.filename)[0]; nome = arquivo.filename
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            arquivo.save(tmp.name); tmp_path = tmp.name
        try:
            extractor = MultiExtractor()
            produtos, info = extractor.extrair_de_pdf(tmp_path, fornecedor)
        finally:
            try: _os.unlink(tmp_path)
            except: pass
        if not produtos:
            return jsonify({'ok': False, 'erro': f'Nenhum produto encontrado em {nome}'})
        conn = get_conn(); cur = conn.cursor(); p = ph(); salvos = 0
        for prod in produtos:
            try:
                cur.execute(f"""INSERT INTO produtos (tenant_id, codigo, fornecedor, descricao, custo, data_analise, arquivo_origem)
                    VALUES ({p},{p},{p},{p},{p},{p},{p})""",
                    [tenant, prod.codigo, fornecedor, prod.descricao, prod.preco_unitario,
                     datetime.now().isoformat(), nome])
                salvos += 1
            except: pass
        conn.commit(); conn.close()
        return jsonify({'ok': True, 'arquivo': nome, 'fornecedor': fornecedor,
                       'produtos_extraidos': salvos, 'provider': info.get('provider','?')})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/upload-pdf', methods=['POST'])
@login_required
def api_upload_pdf():
    import tempfile, os as _os
    if 'arquivo' not in request.files:
        return jsonify({'ok': False, 'erro': 'Nenhum arquivo enviado'})
    arquivo = request.files['arquivo']
    if not arquivo.filename.lower().endswith('.pdf'):
        return jsonify({'ok': False, 'erro': 'Apenas arquivos PDF'})
    try:
        from multi_extractor import MultiExtractor
        tenant = get_tenant_id(); fornecedor = _os.path.splitext(arquivo.filename)[0]; nome = arquivo.filename
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            arquivo.save(tmp.name); tmp_path = tmp.name
        try:
            extractor = MultiExtractor()
            produtos, info = extractor.extrair_de_pdf(tmp_path, fornecedor)
        finally:
            try: _os.unlink(tmp_path)
            except: pass
        if not produtos: return jsonify({'ok': False, 'erro': f'Nenhum produto encontrado em {nome}'})
        conn = get_conn(); cur = conn.cursor(); p = ph(); salvos = 0
        for prod in produtos:
            try:
                cur.execute(f"""INSERT INTO produtos (tenant_id, codigo, fornecedor, descricao, custo, data_analise, arquivo_origem)
                    VALUES ({p},{p},{p},{p},{p},{p},{p})""",
                    [tenant, prod.codigo, fornecedor, prod.descricao, prod.preco_unitario,
                     datetime.now().isoformat(), nome])
                salvos += 1
            except: pass
        conn.commit(); conn.close()
        return jsonify({'ok': True, 'arquivo': nome, 'fornecedor': fornecedor,
                       'produtos_extraidos': salvos, 'provider': info.get('provider','?')})
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/config-pasta', methods=['POST'])
@login_required
def api_config_pasta():
    d = request.get_json(); pasta = d.get('pasta','')
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        if USAR_POSTGRES:
            cur.execute(f"INSERT INTO config (chave,valor) VALUES ({p},{p}) ON CONFLICT (chave) DO UPDATE SET valor={p}",
                       ['pasta_monitorada', pasta, pasta])
        else:
            cur.execute(f"INSERT OR REPLACE INTO config (chave,valor) VALUES ({p},{p})", ['pasta_monitorada', pasta])
        conn.commit(); conn.close()
        return jsonify({'ok': True})
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/verificar-pdfs')
@login_required
def api_verificar_pdfs():
    try:
        conn = get_conn(); cur = conn.cursor(); p = ph()
        cur.execute(f"SELECT valor FROM config WHERE chave = {p}", ['pasta_monitorada'])
        r = cur.fetchone(); conn.close()
        pasta = r[0] if r else ''
        if not pasta: return jsonify({'ok': False, 'erro': 'Pasta não configurada'})
        from pathlib import Path as P
        path = P(pasta)
        if not path.exists(): return jsonify({'ok': False, 'erro': f'Pasta não encontrada: {pasta}'})
        pdfs = list(path.glob('*.pdf'))
        enviados = list(path.glob('*.enviado'))
        return jsonify({'ok': True, 'pasta': pasta, 'pdfs': len(pdfs),
                       'enviados': len(enviados), 'arquivos': [f.name for f in pdfs[:20]]})
    except Exception as e: return jsonify({'ok': False, 'erro': str(e)})

# ════════════════════════════════════════════════════════════════════
# SCHEMA CONFIG TABLE
# ════════════════════════════════════════════════════════════════════
def garantir_config():
    try:
        conn = get_conn(); cur = conn.cursor()
        if USAR_POSTGRES:
            cur.execute("CREATE TABLE IF NOT EXISTS config (chave TEXT PRIMARY KEY, valor TEXT DEFAULT '')")
        else:
            cur.execute("CREATE TABLE IF NOT EXISTS config (chave TEXT PRIMARY KEY, valor TEXT DEFAULT '')")
        conn.commit(); conn.close()
    except: pass


# ════════════════════════════════════════════════════════════════════
# HTML — DASHBOARD PRINCIPAL
# ════════════════════════════════════════════════════════════════════
HTML_DASH = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ brand }} · {{ tenant_nome }}</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
:root{--bg:#0a0e27;--card:#1a1f3a;--border:#2d3452;--text:#e4e6eb;--muted:#8b92a5;
  --green:#4ade80;--red:#f87171;--yellow:#fbbf24;--blue:#667eea;--purple:#c084fc}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--text);font-size:.85rem}
.topbar{background:var(--card);border-bottom:1px solid var(--border);padding:8px 16px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;position:sticky;top:0;z-index:100}
.logo{font-size:1.1rem;font-weight:900;background:linear-gradient(135deg,#667eea,#764ba2);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-right:8px}
.btn{padding:5px 12px;border:none;border-radius:5px;cursor:pointer;font-size:.78rem;font-weight:600;transition:opacity .15s}
.btn:hover{opacity:.85}.btn-blue{background:#667eea;color:#fff}.btn-green{background:#059669;color:#fff}
.btn-yellow{background:#f59e0b;color:#000}.btn-pink{background:#ec4899;color:#fff}
.btn-cyan{background:#0891b2;color:#fff}.btn-purple{background:#7c3aed;color:#fff}
.btn-gray{background:#2d3452;color:var(--text)}.btn-red{background:#dc2626;color:#fff}
.spacer{flex:1}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:8px;padding:12px 16px}
.stat{background:var(--card);border-radius:8px;padding:12px;text-align:center;border:1px solid var(--border)}
.stat-n{font-size:1.5rem;font-weight:800;margin-bottom:2px}
.stat-l{font-size:.65rem;color:var(--muted);text-transform:uppercase;letter-spacing:.5px}
.filters{background:var(--card);border-bottom:1px solid var(--border);padding:10px 16px;display:flex;gap:8px;flex-wrap:wrap;align-items:center}
.filters input,.filters select{background:var(--bg);border:1px solid var(--border);color:var(--text);padding:5px 8px;border-radius:5px;font-size:.78rem}
.filters input:focus,.filters select:focus{border-color:var(--blue);outline:none}
.aliquota-bar{background:var(--card);border-bottom:1px solid var(--border);padding:6px 16px;display:flex;align-items:center;gap:10px;font-size:.78rem;color:var(--muted)}
.aliquota-bar input{width:60px;background:var(--bg);border:1px solid var(--border);color:var(--text);padding:3px 6px;border-radius:4px;text-align:center}
.table-wrap{overflow-x:auto;padding:0 16px 80px}
table{width:100%;border-collapse:collapse;background:var(--card);border-radius:8px;overflow:hidden;font-size:.78rem}
thead{background:#0f1328}
th{padding:8px 6px;color:var(--muted);font-weight:600;text-align:left;white-space:nowrap;border-bottom:1px solid var(--border)}
td{padding:5px 6px;border-bottom:1px solid #151929;vertical-align:middle}
tr:hover td{background:#1f2544}
.badge-ok{background:#064e3b;color:var(--green);padding:1px 6px;border-radius:8px;font-size:.65rem;font-weight:700}
.badge-no{background:#450a0a;color:var(--red);padding:1px 6px;border-radius:8px;font-size:.65rem;font-weight:700}
.badge-pend{background:#3b2700;color:var(--yellow);padding:1px 6px;border-radius:8px;font-size:.65rem;font-weight:700}
.inline-input{background:transparent;border:none;color:var(--text);width:70px;text-align:right;font-size:.78rem;padding:2px 4px;border-radius:3px}
.inline-input:focus{background:#2d3452;outline:none}
.inline-input:hover{background:#2d3452}
.pagination{display:flex;gap:6px;justify-content:center;padding:12px}
.pg-btn{background:var(--card);border:1px solid var(--border);color:var(--text);padding:4px 10px;border-radius:4px;cursor:pointer;font-size:.75rem;text-decoration:none}
.pg-btn.active{background:var(--blue);border-color:var(--blue);font-weight:700}
.modal-bg{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.8);display:flex;justify-content:center;align-items:center;z-index:9999;padding:20px}
.modal{background:var(--card);padding:20px;border-radius:10px;width:500px;max-width:95vw;max-height:90vh;overflow-y:auto}
.modal h3{margin-bottom:14px;font-size:1rem}
.modal label{display:block;color:var(--muted);font-size:.72rem;text-transform:uppercase;margin-bottom:4px}
.modal input,.modal select{width:100%;background:var(--bg);border:1px solid var(--border);color:var(--text);padding:8px 10px;border-radius:5px;margin-bottom:12px;font-size:.85rem}
.toast{position:fixed;bottom:20px;right:20px;padding:10px 16px;border-radius:6px;font-size:.82rem;font-weight:600;z-index:9999;animation:fadeIn .2s}
@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
</style>
</head>
<body>

<div class="topbar">
  <span class="logo" title="{{ plataforma }} · {{ tenant_nome }}">{{ brand }}</span>
  <span style="color:var(--muted);font-size:.72rem"><span style="color:#c084fc;font-weight:600">{{ tenant_nome }}</span> · {{ usuario }}{% if is_admin %} · <a href="/admin/usuarios" style="color:#fbbf24;text-decoration:none">⚙️</a>{% endif %}</span>
  <div class="spacer"></div>
  <button class="btn btn-yellow" onclick="abrirAlertaDiario()">📊 Alerta</button>
  <button class="btn btn-pink" onclick="abrirTendencias()">🔥 Tendências</button>
  <button class="btn btn-cyan" onclick="abrirSaude()">🏥 Saúde</button>
  <a href="/ml" class="btn" style="background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;text-decoration:none">🚀 Painel ML</a>
  <button class="btn btn-green" onclick="mostrarModalProduto()">➕ Produto</button>
  <button class="btn btn-blue" onclick="exportar()">📥 Excel</button>
  <button class="btn btn-gray" onclick="window.location='/escolhidos'">⭐ Escolhidos ({{ stats.escolhidos }})</button>
  <button class="btn btn-gray" onclick="location.href='/config'">⚙️ Config</button>
  <button class="btn btn-gray" onclick="location.href='/logout'">Sair</button>
</div>

<div class="aliquota-bar">
  <span>🏛️ Alíquota Imposto (%):</span>
  <input type="number" id="aliquota" value="{{ imp_pct }}" min="0" max="100" step="0.5" onchange="salvarAliquota(this.value)">
  <span style="font-size:.7rem">Aplica a todos ao recalcular</span>
  <button class="btn btn-gray" style="padding:3px 8px;font-size:.7rem" onclick="recalcularTodos()">⚡ Recalcular Tudo</button>
</div>

<div class="stats">
  <div class="stat"><div class="stat-n" style="color:var(--blue)">{{ stats.total }}</div><div class="stat-l">Total</div></div>
  <div class="stat"><div class="stat-n" style="color:var(--muted)">{{ stats.fornecedores }}</div><div class="stat-l">Fornecedores</div></div>
  <div class="stat"><div class="stat-n" style="color:var(--green)">{{ stats.viaveis }}</div><div class="stat-l">✅ Viáveis ≥20%</div></div>
  <div class="stat"><div class="stat-n" style="color:var(--red)">{{ stats.nao_viaveis }}</div><div class="stat-l">❌ Não Viáveis</div></div>
  <div class="stat"><div class="stat-n" style="color:var(--yellow)">{{ stats.pendentes }}</div><div class="stat-l">⏳ Pendentes</div></div>
  <div class="stat"><div class="stat-n" style="color:var(--purple)">{{ stats.escolhidos }}</div><div class="stat-l">⭐ Escolhidos</div></div>
</div>

<div class="filters">
  <span style="color:var(--muted);font-size:.72rem">Filtros:</span>
  <input type="text" placeholder="Fornecedor" id="ff" value="{{ f.forn }}" onkeydown="if(event.key=='Enter')filtrar()">
  <input type="text" placeholder="Produto" id="fp" value="{{ f.produto }}" onkeydown="if(event.key=='Enter')filtrar()">
  <select id="fs" onchange="filtrar()">
    <option value="" {% if f.status=='' %}selected{% endif %}>Todos Status</option>
    <option value="viavel" {% if f.status=='viavel' %}selected{% endif %}>✅ Viáveis</option>
    <option value="nao_viavel" {% if f.status=='nao_viavel' %}selected{% endif %}>❌ Não Viáveis</option>
    <option value="pendente" {% if f.status=='pendente' %}selected{% endif %}>⏳ Pendentes</option>
    <option value="escolhido" {% if f.status=='escolhido' %}selected{% endif %}>⭐ Escolhidos</option>
  </select>
  <input type="number" placeholder="Custo mín" id="fcmin" value="{{ f.cmin }}" style="width:90px" onkeydown="if(event.key=='Enter')filtrar()">
  <input type="number" placeholder="Custo máx" id="fcmax" value="{{ f.cmax }}" style="width:90px" onkeydown="if(event.key=='Enter')filtrar()">
  <select id="fpp" onchange="filtrar()">
    <option value="50" {% if f.pp==50 %}selected{% endif %}>50/pág</option>
    <option value="100" {% if f.pp==100 %}selected{% endif %}>100/pág</option>
    <option value="200" {% if f.pp==200 %}selected{% endif %}>200/pág</option>
  </select>
  <button class="btn btn-blue" onclick="filtrar()">🔍</button>
  <button class="btn btn-gray" onclick="limparFiltros()">✕</button>
  <span style="color:var(--muted);font-size:.72rem;margin-left:auto">{{ total_filtrado }} produto(s)</span>
</div>

<div class="table-wrap">
<table>
<thead><tr>
  <th style="width:30px"></th>
  <th>Fornecedor</th><th>Código</th><th style="min-width:200px">Descrição</th>
  <th>Custo</th><th>Preço ML</th><th>Taxa%</th><th>Frete</th>
  <th>Custo Total</th><th>Margem%</th><th>Margem R$</th><th>Status</th>
  <th>Ações</th>
</tr></thead>
<tbody>
{% for p in produtos %}
<tr>
  <td>
    {% if p.escolhido %}
    <button title="Descartar" onclick="descartar({{ p.id }}, this)" style="background:none;border:none;cursor:pointer;font-size:.9rem">⭐</button>
    {% else %}
    <button title="Escolher" onclick="escolher({{ p.id }}, this)" style="background:none;border:none;cursor:pointer;font-size:.9rem;opacity:.4">☆</button>
    {% endif %}
  </td>
  <td style="max-width:100px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--muted)" title="{{ p.fornecedor }}">{{ p.fornecedor }}</td>
  <td style="color:var(--muted)">{{ p.codigo or '' }}</td>
  <td style="max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="{{ p.descricao }}">{{ p.descricao }}</td>
  <td style="color:var(--yellow)">R$ {{ p.custo|br }}</td>
  <td><input class="inline-input" id="pml-{{ p.id }}" value="{{ p.preco_ml or '' }}" placeholder="0,00" onblur="salvar(this,'preco_ml',{{ p.id }})" onkeydown="if(event.key=='Enter')this.blur()"></td>
  <td><input class="inline-input" id="tx-{{ p.id }}" value="{{ ((p.taxa_categoria or 0.165)*100)|br(1) }}" style="width:50px" onblur="salvarTaxa(this,{{ p.id }})" onkeydown="if(event.key=='Enter')this.blur()"></td>
  <td style="color:var(--muted)">{{ p.custo_frete|br if p.custo_frete else '-' }}</td>
  <td>{{ p.custo_total|br if p.custo_total else '-' }}</td>
  <td style="color:{{ '#4ade80' if (p.margem_percentual or 0) >= 20 else '#f87171' if p.preco_ml else '#8b92a5' }};font-weight:700">
    {{ p.margem_percentual|br(1) if p.preco_ml else '-' }}%
  </td>
  <td style="color:{{ '#4ade80' if (p.margem_reais or 0) >= 0 else '#f87171' }}">
    {{ p.margem_reais|br if p.preco_ml else '-' }}
  </td>
  <td>
    {% if not p.preco_ml %}<span class="badge-pend">PENDENTE</span>
    {% elif p.viavel %}<span class="badge-ok">VIÁVEL</span>
    {% else %}<span class="badge-no">BAIXO</span>{% endif %}
  </td>
  <td style="display:flex;gap:3px;flex-wrap:nowrap">
    <button class="btn btn-gray" style="padding:2px 6px;font-size:.7rem" 
            onclick="buscarML({{ p.id }}, '{{ p.descricao|replace("'", "\\'") }}')" title="Buscar preço ML">🔍</button>
    <button class="btn btn-purple" style="padding:2px 6px;font-size:.7rem"
            onclick="pesquisarProduto({{ p.id }})" title="Agente: Pesquisa">🤖</button>
    <button class="btn btn-green" style="padding:2px 6px;font-size:.7rem"
            data-id="{{ p.id }}" data-desc="{{ p.descricao|e }}" data-custo="{{ p.custo or 0 }}" data-peso="{{ p.peso_kg or 0 }}"
            onclick="precificar(this)" title="Agente: Precificação">💰</button>
    <button class="btn" style="padding:2px 6px;font-size:.7rem;background:#0891b2;color:#fff"
            data-id="{{ p.id }}" data-desc="{{ p.descricao|e }}" data-custo="{{ p.custo or 0 }}" data-peso="{{ p.peso_kg or 0 }}"
            onclick="viabilidade(this)" title="Agente: Viabilidade">📈</button>
    <button class="btn btn-red" style="padding:2px 6px;font-size:.7rem"
            onclick="deletar({{ p.id }})" title="Deletar">🗑️</button>
  </td>
</tr>
{% else %}
<tr><td colspan="13" style="text-align:center;padding:40px;color:var(--muted)">
  📭 Nenhum produto encontrado. <a href="/config" style="color:var(--blue)">Envie PDFs em Configurações →</a>
</td></tr>
{% endfor %}
</tbody>
</table>
</div>

{% if total_paginas > 1 %}
<div class="pagination">
  {% if pg > 1 %}<a class="pg-btn" href="{{ url_pg(pg-1) }}">← Anterior</a>{% endif %}
  {% for i in range([1,pg-2]|max, [total_paginas+1,pg+3]|min) %}
    <a class="pg-btn {% if i==pg %}active{% endif %}" href="{{ url_pg(i) }}">{{ i }}</a>
  {% endfor %}
  {% if pg < total_paginas %}<a class="pg-btn" href="{{ url_pg(pg+1) }}">Próxima →</a>{% endif %}
</div>
{% endif %}

<!-- Modal Produto -->
<div id="modalProd" style="display:none" class="modal-bg">
<div class="modal" style="border:1px solid var(--green)">
  <h3 style="color:var(--green)">➕ Novo Produto</h3>
  <label>Fornecedor</label><input id="np-forn" placeholder="Ex: Fornecedor ABC">
  <label>Descrição</label><input id="np-desc" placeholder="Nome do produto">
  <label>Custo (R$)</label><input id="np-custo" type="number" placeholder="0,00">
  <label>Código</label><input id="np-cod" placeholder="Código interno (opcional)">
  <div style="display:flex;gap:8px">
    <button class="btn btn-green" style="flex:1" onclick="adicionarProduto()">Adicionar</button>
    <button class="btn btn-gray" style="flex:1" onclick="document.getElementById('modalProd').style.display='none'">Cancelar</button>
  </div>
</div>
</div>

<script>
const fmt = (v,d=2) => v!=null ? parseFloat(v).toFixed(d).replace('.',',') : '-';

function showToast(msg, erro=false){
  document.querySelectorAll('.toast').forEach(t=>t.remove());
  const t = document.createElement('div');
  t.className='toast'; t.textContent=msg;
  t.style.background = erro ? '#450a0a' : '#064e3b';
  t.style.color = erro ? '#f87171' : '#4ade80';
  t.style.border = `1px solid ${erro?'#f87171':'#4ade80'}`;
  document.body.appendChild(t);
  setTimeout(()=>t.remove(), 3500);
}

function filtrar(){
  const p = new URLSearchParams();
  const forn = document.getElementById('ff').value;
  const prod = document.getElementById('fp').value;
  const st = document.getElementById('fs').value;
  const cmin = document.getElementById('fcmin').value;
  const cmax = document.getElementById('fcmax').value;
  const pp = document.getElementById('fpp').value;
  if(forn) p.set('forn',forn); if(prod) p.set('produto',prod);
  if(st) p.set('status',st); if(cmin) p.set('cmin',cmin);
  if(cmax && cmax!='999') p.set('cmax',cmax); if(pp!='100') p.set('pp',pp);
  window.location = '/?' + p.toString();
}

function limparFiltros(){ window.location = '/'; }
function exportar(){ window.open('/exportar','_blank'); }

function salvar(el, campo, id){
  let val = el.value.replace(',','.');
  if(!val || isNaN(parseFloat(val))) return;
  fetch('/api/atualizar',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({id,campo,valor:parseFloat(val)})})
  .then(r=>r.json()).then(d=>{ if(d.ok) setTimeout(()=>location.reload(),300); else showToast('❌ '+d.erro,true); })
  .catch(()=>showToast('❌ Erro',true));
}

function salvarTaxa(el, id){
  let val = parseFloat(el.value.replace(',','.')) / 100;
  if(isNaN(val)) return;
  fetch('/api/atualizar',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({id,campo:'taxa_categoria',valor:val})})
  .then(r=>r.json()).then(d=>{ if(d.ok) setTimeout(()=>location.reload(),300); });
}

function salvarAliquota(v){
  fetch('/api/aliquota',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({valor:parseFloat(v)||0})}).then(r=>r.json())
  .then(d=>{ if(d.ok) showToast('✅ Alíquota salva'); });
}

function recalcularTodos(){
  showToast('⚡ Recalculando...');
  fetch('/api/recalcular-todos',{method:'POST'}).then(r=>r.json())
  .then(d=>{ showToast('✅ '+d.atualizados+' produtos recalculados'); setTimeout(()=>location.reload(),800); })
  .catch(()=>showToast('❌ Erro',true));
}

function escolher(id, btn){
  fetch('/api/escolher',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})})
  .then(r=>r.json()).then(d=>{ if(d.ok){btn.textContent='⭐';btn.style.opacity='1';showToast('⭐ Escolhido!'); } });
}
function descartar(id, btn){
  fetch('/api/descartar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})})
  .then(r=>r.json()).then(d=>{ if(d.ok){btn.textContent='☆';btn.style.opacity='.4'; } });
}

function deletar(id){
  if(!confirm('Deletar este produto?')) return;
  fetch('/api/deletar-produto',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})})
  .then(r=>r.json()).then(d=>{ if(d.ok){showToast('🗑️ Deletado');setTimeout(()=>location.reload(),500);}else showToast('❌ '+d.erro,true); });
}

function mostrarModalProduto(){ document.getElementById('modalProd').style.display='flex'; document.getElementById('np-forn').focus(); }

function adicionarProduto(){
  const forn=document.getElementById('np-forn').value.trim();
  const desc=document.getElementById('np-desc').value.trim();
  const custo=parseFloat(document.getElementById('np-custo').value||0);
  const cod=document.getElementById('np-cod').value.trim();
  if(!desc){showToast('❌ Descrição obrigatória',true);return;}
  fetch('/api/adicionar-produto',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({fornecedor:forn,descricao:desc,custo,codigo:cod})})
  .then(r=>r.json()).then(d=>{
    if(d.ok){document.getElementById('modalProd').style.display='none';showToast('✅ '+d.desc+' adicionado!');setTimeout(()=>location.reload(),600);}
    else showToast('❌ '+d.erro,true);
  });
}

function buscarML(id, desc){
  const btn = event.target;
  btn.textContent='⏳'; btn.disabled=true;
  fetch('/api/buscar-ml',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id,termo:desc})})
  .then(r=>r.json()).then(d=>{
    btn.textContent='🔍'; btn.disabled=false;
    if(d.ok){
      const el=document.getElementById('pml-'+id);
      const tx=document.getElementById('tx-'+id);
      if(el && !el.value){ el.value=fmt(d.preco_medio); }
      if(tx && d.taxa_percentual) tx.value=fmt(d.taxa_percentual,1);
      showToast('ML: R$ '+fmt(d.preco_medio)+' | Taxa: '+fmt(d.taxa_percentual,1)+'% ('+d.total+' res.)');
    } else showToast('❌ '+d.erro,true);
  }).catch(()=>{btn.textContent='🔍';btn.disabled=false;showToast('❌ Erro',true);});
}

// ── Agentes IA ────────────────────────────────────────────────────
function pesquisarProduto(id){
  showToast('🤖 Pesquisando...');
  fetch('/api/pesquisar-produto',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})})
  .then(r=>r.json()).then(d=>{ if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalAgente('🤖 Pesquisa de Produto',d); })
  .catch(()=>showToast('❌ Erro',true));
}

function precificar(btn){
  const id=btn.dataset.id; const desc=btn.dataset.desc; const custo=btn.dataset.custo; const peso=btn.dataset.peso;
  const margem=prompt('Margem mínima (%):', '20'); if(margem===null)return;
  showToast('💰 Calculando precificação...');
  fetch('/api/precificar',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({id:parseInt(id),margem_minima:parseFloat(margem)||20})})
  .then(r=>r.json()).then(d=>{ if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalPrecificacao(d); })
  .catch(()=>showToast('❌ Erro',true));
}

function viabilidade(btn){
  const id=btn.dataset.id; const desc=btn.dataset.desc; const custo=parseFloat(btn.dataset.custo)||0; const peso=parseFloat(btn.dataset.peso)||0;
  const margem=prompt('Margem mínima (%):', '20'); if(margem===null)return;
  showToast('📈 Analisando viabilidade...');
  fetch('/api/viabilidade',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({id:parseInt(id),termo:desc,custo,peso,margem_minima:parseFloat(margem)||20})})
  .then(r=>r.json()).then(d=>{ if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalViabilidade(d); })
  .catch(()=>showToast('❌ Erro',true));
}

function abrirAlertaDiario(){
  showToast('📊 Gerando alerta...');
  fetch('/api/alerta-diario').then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalAlerta(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function abrirTendencias(){
  showToast('🔥 Buscando tendências...');
  fetch('/api/tendencias').then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalTendencias(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function abrirSaude(){
  showToast('🏥 Verificando anúncios...');
  fetch('/api/saude-anuncios').then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalSaude(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function mostrarModalAgente(titulo, d){
  const h = `<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #7c3aed;max-width:600px" onclick="event.stopPropagation()">
    <h3 style="color:#c084fc;margin-bottom:12px">${titulo}</h3>
    <pre style="background:#0a0e27;padding:12px;border-radius:6px;overflow-x:auto;font-size:.75rem;color:#e4e6eb;white-space:pre-wrap">${JSON.stringify(d,null,2)}</pre>
    <button class="btn btn-gray" style="margin-top:10px;width:100%" onclick="this.closest('.modal-bg').remove()">Fechar</button>
  </div></div>`;
  document.body.insertAdjacentHTML('beforeend',h);
}

function mostrarModalPrecificacao(d){
  const fmt2=(v,dc=2)=>v!=null?parseFloat(v).toFixed(dc).replace('.',','):'-';
  const c=d.cenarios||{}; const rec=d.recomendacao||{};
  const renderC=(nome,dados,emoji)=>{
    if(!dados)return'';
    const dest=rec.cenario===nome?'border:2px solid #667eea;':'';
    const recTag=rec.cenario===nome?'<span style="background:#667eea;color:#fff;padding:1px 5px;border-radius:8px;font-size:.6rem;margin-left:4px">⭐</span>':'';
    const cor=dados.viavel?'#4ade80':'#f87171';
    return `<div style="background:#0a0e27;padding:10px;border-radius:6px;${dest}">
      <div style="display:flex;justify-content:space-between;margin-bottom:4px">
        <span style="font-size:.8rem;font-weight:700">${emoji} ${nome}${recTag}</span>
        <span style="color:${cor};font-size:.65rem;font-weight:700">${dados.viavel?'VIÁVEL':'INVIÁVEL'}</span>
      </div>
      <div style="font-size:1.2rem;font-weight:700;color:#4ade80">R$ ${fmt2(dados.preco)}</div>
      <div style="color:${cor};font-size:.8rem">Margem: ${fmt2(dados.margem,1)}%</div>
      <div style="color:#8b92a5;font-size:.68rem;margin-top:3px">${dados.descricao}</div>
      <button onclick="aplicarPreco(${d.produto_id},${dados.preco},${d.taxa_percentual})"
        style="margin-top:6px;padding:2px 8px;border:none;border-radius:3px;background:#2d3452;color:#e4e6eb;cursor:pointer;font-size:.7rem">Usar</button>
    </div>`;
  };
  const rivais=(d.top_concorrentes||[]).map((r,i)=>`<tr style="border-bottom:1px solid #1a1f3a">
    <td style="padding:3px 6px;color:#8b92a5">${i+1}</td>
    <td style="padding:3px 6px;color:#e4e6eb;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${r.titulo}</td>
    <td style="padding:3px 6px;color:#4ade80">R$ ${fmt2(r.preco)}</td>
    <td style="padding:3px 6px;color:#fbbf24">${parseInt(r.vendas_acumuladas||0).toLocaleString('pt-BR')}</td>
    <td><a href="${r.link}" target="_blank" style="color:#667eea;font-size:.7rem">↗</a></td>
  </tr>`).join('');
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #059669;width:820px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;margin-bottom:10px">
      <div><h3 style="color:#059669">💰 Precificação Inteligente</h3>
      <div style="color:#8b92a5;font-size:.72rem">${d.produto_nome} · Taxa: ${fmt2(d.taxa_percentual,1)}% · "${d.termo_buscado}"</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Menor Rival</div><div style="font-size:1.1rem;font-weight:700;color:#f87171">R$ ${fmt2(d.preco_min_ml)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Mediano</div><div style="font-size:1.1rem;font-weight:700;color:#4ade80">R$ ${fmt2(d.preco_mediano)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Seu Custo</div><div style="font-size:1.1rem;font-weight:700;color:#fbbf24">R$ ${fmt2(d.custo)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Mín Viável</div><div style="font-size:1.1rem;font-weight:700;color:#c084fc">R$ ${fmt2(d.preco_minimo_viavel)}</div></div>
    </div>
    ${rec.motivo?`<div style="background:#1a2a1a;border:1px solid #059669;border-radius:5px;padding:10px;margin-bottom:12px;font-size:.82rem;color:#e4e6eb">💡 ${rec.motivo}</div>`:''}
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px">
      ${renderC('agressivo',c.agressivo,'🔥')}
      ${renderC('competitivo',c.competitivo,'⚖️')}
      ${renderC('premium',c.premium,'👑')}
    </div>
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.75rem">
      <thead><tr style="border-bottom:1px solid #2d3452"><th style="padding:3px 6px;color:#8b92a5">#</th><th style="padding:3px 6px;color:#8b92a5">Título</th><th style="padding:3px 6px;color:#8b92a5">Preço</th><th style="padding:3px 6px;color:#8b92a5">Vendas</th><th></th></tr></thead>
      <tbody>${rivais}</tbody>
    </table>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function aplicarPreco(id,preco,taxa){
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  const el=document.getElementById('pml-'+id); const tx=document.getElementById('tx-'+id);
  if(el){ el.value=preco.toFixed(2); if(tx&&taxa) tx.value=taxa.toFixed(1); salvar(el,'preco_ml',id); showToast('✅ Preço aplicado!'); }
}

function mostrarModalViabilidade(d){
  const fmt2=(v,dc=2)=>v!=null?parseFloat(v).toFixed(dc).replace('.',','):'-';
  const corS=d.score>=70?'#4ade80':d.score>=45?'#fbbf24':'#f87171';
  const ins=(d.insights||[]).map(i=>`<div style="display:flex;gap:8px;padding:5px 0;border-bottom:1px solid #2d3452;font-size:.8rem">
    <span>${i.tipo==='positivo'?'✅':i.tipo==='negativo'?'❌':'⚠️'}</span>
    <span style="color:#e4e6eb">${i.msg}</span></div>`).join('');
  const top=(d.top_anuncios||[]).map((a,i)=>`<tr style="border-bottom:1px solid #1a1f3a">
    <td style="padding:3px 6px;color:#8b92a5">${i+1}</td>
    <td style="padding:3px 6px;color:#e4e6eb;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${a.titulo}</td>
    <td style="padding:3px 6px;color:#4ade80">R$ ${fmt2(a.preco)}</td>
    <td style="padding:3px 6px;color:#fbbf24">${parseInt(a.vendas_acumuladas||0).toLocaleString('pt-BR')}</td>
    <td><a href="${a.link}" target="_blank" style="color:#667eea;font-size:.7rem">↗</a></td>
  </tr>`).join('');
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #0891b2;width:820px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;margin-bottom:12px">
      <div><h3 style="color:#0891b2">📈 Viabilidade de Produto</h3>
      <div style="color:#8b92a5;font-size:.72rem">${d.termo_buscado} · ${d.total_anuncios} anúncios</div></div>
      <div style="text-align:center"><div style="font-size:2rem;font-weight:900;color:${corS}">${d.score}</div><div style="font-size:.7rem;color:${corS}">${d.cor} ${d.classificacao}</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Vendas Top20</div><div style="font-size:1.1rem;font-weight:700;color:#4ade80">${parseInt(d.total_vendas_top20_acumulado||0).toLocaleString('pt-BR')}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Demanda/mês</div><div style="font-size:1.1rem;font-weight:700;color:#fbbf24">${fmt2(d.demanda_mercado_mensal,0)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Preço Mínimo</div><div style="font-size:1.1rem;font-weight:700;color:#f87171">R$ ${fmt2(d.preco_min)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Mediano</div><div style="font-size:1.1rem;font-weight:700;color:#4ade80">R$ ${fmt2(d.preco_mediano)}</div></div>
    </div>
    <div style="background:#0a0e27;border-radius:6px;padding:10px;margin-bottom:12px">${ins}</div>
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.75rem">
      <thead><tr style="border-bottom:1px solid #2d3452"><th style="padding:3px 6px;color:#8b92a5">#</th><th style="padding:3px 6px;color:#8b92a5">Título</th><th style="padding:3px 6px;color:#8b92a5">Preço</th><th style="padding:3px 6px;color:#8b92a5">Vendas</th><th></th></tr></thead>
      <tbody>${top}</tbody>
    </table>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function mostrarModalAlerta(d){
  const vendas=d.vendas||{}; const rep=d.reputacao||{}; const pergs=d.perguntas||{};
  const fmt2=(v)=>v!=null?parseFloat(v).toFixed(2).replace('.',','):'0,00';
  const alertasH=(d.alertas||[]).map(a=>{
    const bg=a.tipo==='urgente'?'#450a0a':a.tipo==='aviso'?'#3b2700':a.tipo==='ok'?'#064e3b':'#1e3a5f';
    const cor=a.tipo==='urgente'?'#f87171':a.tipo==='aviso'?'#fbbf24':a.tipo==='ok'?'#4ade80':'#93c5fd';
    return `<div style="background:${bg};border-radius:5px;padding:8px 12px;display:flex;align-items:center;gap:8px">
      <span>${a.icone}</span><span style="color:${cor};font-size:.82rem">${a.msg}</span></div>`;
  }).join('');
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #f59e0b;width:650px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;margin-bottom:12px">
      <div><h3 style="color:#f59e0b">📊 Alerta Diário</h3><div style="color:#8b92a5;font-size:.72rem">${d.gerado_em}</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="display:flex;flex-direction:column;gap:6px;margin-bottom:12px">${alertasH||'<div style="color:#8b92a5">Sem alertas.</div>'}</div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Vendas Hoje</div><div style="font-size:1.2rem;font-weight:700;color:#4ade80">${vendas.hoje_qtd||0}</div><div style="font-size:.72rem;color:#8b92a5">R$ ${fmt2(vendas.hoje_valor)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Vendas Semana</div><div style="font-size:1.2rem;font-weight:700;color:#667eea">${vendas.semana_qtd||0}</div><div style="font-size:.72rem;color:#8b92a5">R$ ${fmt2(vendas.semana_valor)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Perguntas</div><div style="font-size:1.2rem;font-weight:700;color:${(pergs.nao_respondidas||0)>0?'#f87171':'#4ade80'}">${pergs.nao_respondidas||0}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Reputação</div><div style="font-size:.82rem;font-weight:700;color:#fbbf24">${(rep.nivel||'N/A').replace(/_/g,' ')}</div></div>
    </div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function mostrarModalTendencias(d){
  const fmt2=(v,dc=2)=>v!=null?parseFloat(v).toFixed(dc).replace('.',','):'-';
  const linhas=(d.tendencias||[]).map((t,i)=>{
    const op=t.oportunidade||{};
    const badge=t.ja_tenho_no_catalogo?'<span style="background:#065f46;color:#4ade80;padding:0 5px;border-radius:8px;font-size:.62rem">✅ Tenho</span>':'';
    return `<tr style="border-bottom:1px solid #1a1f3a;${t.ja_tenho_no_catalogo?'background:#0a1a0a':''}">
      <td style="padding:4px 7px;color:#8b92a5">${i+1}</td>
      <td style="padding:4px 7px;color:#e4e6eb;font-weight:600">${t.keyword} ${badge}</td>
      <td style="padding:4px 7px;color:#4ade80">${t.preco_min?'R$ '+fmt2(t.preco_min):'-'}</td>
      <td style="padding:4px 7px;color:#fbbf24">${t.preco_medio?'R$ '+fmt2(t.preco_medio):'-'}</td>
      <td style="padding:4px 7px"><span style="color:${op.cor||'#8b92a5'};font-size:.78rem">${op.icone||''} ${op.label||''}</span></td>
      <td><a href="${t.url_ml||'https://www.mercadolivre.com.br/'+encodeURIComponent(t.keyword)}" target="_blank" style="color:#667eea;font-size:.72rem">↗</a></td>
    </tr>`;
  }).join('');
  const tenhoC=(d.tendencias||[]).filter(t=>t.ja_tenho_no_catalogo).length;
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #ec4899;width:850px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;margin-bottom:12px">
      <div><h3 style="color:#ec4899">🔥 Tendências ML</h3><div style="color:#8b92a5;font-size:.72rem">${d.total} produtos em alta · ${tenhoC} você tem</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.8rem">
      <thead><tr style="border-bottom:1px solid #2d3452"><th style="padding:5px 7px;color:#8b92a5">#</th><th style="padding:5px 7px;color:#8b92a5">Produto</th><th style="padding:5px 7px;color:#8b92a5">Mín</th><th style="padding:5px 7px;color:#8b92a5">Médio</th><th style="padding:5px 7px;color:#8b92a5">Oportunidade</th><th style="padding:5px 7px;color:#8b92a5">ML</th></tr></thead>
      <tbody>${linhas}</tbody>
    </table></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function mostrarModalSaude(d){
  const r=d.resumo||{}; const corS=d.score_saude>=90?'#4ade80':d.score_saude>=70?'#fbbf24':'#f87171';
  const renderL=(lista,titulo,cor,icone)=>{
    if(!lista||!lista.length)return'';
    return `<div style="margin-bottom:10px"><div style="color:${cor};font-weight:700;font-size:.8rem;margin-bottom:4px">${icone} ${titulo} (${lista.length})</div>
      <div style="background:#0a0e27;border-radius:5px">${lista.map(a=>`<div style="display:flex;justify-content:space-between;padding:5px 10px;border-bottom:1px solid #1a1f3a;font-size:.78rem">
        <span style="color:#e4e6eb">${a.titulo||a.id}</span>
        ${a.link?`<a href="${a.link}" target="_blank" style="color:#667eea">↗</a>`:''}
      </div>`).join('')}</div></div>`;
  };
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #0284c7;width:750px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;margin-bottom:12px">
      <div><h3 style="color:#0284c7">🏥 Saúde dos Anúncios</h3><div style="color:#8b92a5;font-size:.72rem">${d.gerado_em}</div></div>
      <div style="text-align:center"><div style="font-size:2rem;font-weight:900;color:${corS}">${d.score_saude}</div><div style="font-size:.72rem;color:${corS}">${d.classificacao_saude}</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Total</div><div style="font-size:1.3rem;font-weight:700;color:#667eea">${r.total_anuncios||0}</div></div>
      <div style="background:#064e3b;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Saudáveis</div><div style="font-size:1.3rem;font-weight:700;color:#4ade80">${r.total_saudaveis||0}</div></div>
      <div style="background:#3b2700;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Em Risco</div><div style="font-size:1.3rem;font-weight:700;color:#fbbf24">${r.total_warning||0}</div></div>
      <div style="background:#450a0a;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Críticos</div><div style="font-size:1.3rem;font-weight:700;color:#f87171">${r.total_unhealthy||0}</div></div>
    </div>
    ${renderL(d.unhealthy,'Críticos — Perdendo Exposição','#f87171','🚨')}
    ${renderL(d.warning,'Em Risco','#fbbf24','⚠️')}
    ${renderL(d.sem_visitas,'Sem Visitas (30 dias)','#8b92a5','👻')}
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

// ── ML Painel (em /ml) ────────────────────────────────────────────
// As funcionalidades de ML (Pedidos, Perguntas, Anúncios, Reputação,
// Faturamento, Métricas, Price to Win, Webhooks, Publicar) foram
// movidas para a rota /ml (página separada).

// Atalho Enter no filtro
document.addEventListener('keydown', e => { if(e.key === 'Escape') document.querySelectorAll('.modal-bg').forEach(m=>m.remove()); });
</script>
</body></html>"""


# ════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════
# HTML — PAINEL ML (PÁGINA SEPARADA)
# ════════════════════════════════════════════════════════════════════
HTML_ML = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ brand }} — Painel ML</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
:root{--bg:#0a0e27;--card:#1a1f3a;--border:#2d3452;--text:#e4e6eb;--muted:#8b92a5;
  --green:#4ade80;--red:#f87171;--yellow:#fbbf24;--blue:#667eea;--purple:#c084fc}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--text);font-size:.85rem}
.topbar{background:var(--card);border-bottom:1px solid var(--border);padding:10px 16px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;position:sticky;top:0;z-index:100}
.logo{font-size:1.1rem;font-weight:900;background:linear-gradient(135deg,#667eea,#764ba2);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-right:8px}
.btn{padding:5px 12px;border:none;border-radius:5px;cursor:pointer;font-size:.78rem;font-weight:600;transition:opacity .15s;text-decoration:none;display:inline-block}
.btn:hover{opacity:.85}.btn-blue{background:#667eea;color:#fff}.btn-green{background:#059669;color:#fff}
.btn-yellow{background:#f59e0b;color:#000}.btn-pink{background:#ec4899;color:#fff}
.btn-cyan{background:#0891b2;color:#fff}.btn-purple{background:#7c3aed;color:#fff}
.btn-gray{background:#2d3452;color:var(--text)}.btn-red{background:#dc2626;color:#fff}
.spacer{flex:1}
.hero{padding:22px 20px 8px;text-align:center}
.hero h1{font-size:1.4rem;font-weight:800;background:linear-gradient(135deg,#667eea,#764ba2);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:4px}
.hero p{color:var(--muted);font-size:.82rem}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:14px;padding:18px 20px 60px;max-width:1280px;margin:0 auto}
.card{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:18px;cursor:pointer;transition:transform .15s,border-color .15s,box-shadow .15s}
.card:hover{transform:translateY(-3px);border-color:var(--blue);box-shadow:0 8px 24px rgba(102,126,234,.15)}
.card-ico{font-size:1.9rem;margin-bottom:8px;display:block}
.card-title{font-size:.95rem;font-weight:700;margin-bottom:4px;color:var(--text)}
.card-desc{color:var(--muted);font-size:.75rem;line-height:1.45}
.card-tag{display:inline-block;background:#0a0e27;border:1px solid var(--border);color:var(--muted);font-size:.6rem;padding:2px 7px;border-radius:10px;margin-top:10px;font-family:monospace}
.modal-bg{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.8);display:flex;justify-content:center;align-items:center;z-index:9999;padding:20px}
.modal{background:var(--card);padding:20px;border-radius:10px;width:500px;max-width:95vw;max-height:90vh;overflow-y:auto}
.modal h3{margin-bottom:14px;font-size:1rem}
.modal label{display:block;color:var(--muted);font-size:.72rem;text-transform:uppercase;margin-bottom:4px}
.modal input,.modal select{width:100%;background:var(--bg);border:1px solid var(--border);color:var(--text);padding:8px 10px;border-radius:5px;margin-bottom:12px;font-size:.85rem}
.toast{position:fixed;bottom:20px;right:20px;padding:10px 16px;border-radius:6px;font-size:.82rem;font-weight:600;z-index:9999;animation:fadeIn .2s}
@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
</style>
</head>
<body>

<div class="topbar">
  <span class="logo" title="{{ plataforma }} · {{ tenant_nome }}">{{ brand }}</span>
  <span style="color:var(--muted);font-size:.72rem"><span style="color:#c084fc;font-weight:600">{{ tenant_nome }}</span> · {{ usuario }}{% if is_admin %} · <a href="/admin/usuarios" style="color:#fbbf24;text-decoration:none">⚙️</a>{% endif %}</span>
  <div class="spacer"></div>
  <a href="/" class="btn btn-gray">← Dashboard</a>
  <a href="/config" class="btn btn-gray">⚙️ Config</a>
</div>

<div class="hero">
  <h1>🚀 Painel ML — Integração Mercado Livre</h1>
  <p>Todas as métricas e operações via API oficial do ML · <span id="ml-status" style="color:var(--yellow)">verificando conexão...</span></p>
</div>

<div class="grid">
  <div class="card" onclick="abrirPedidos()"><span class="card-ico">🛒</span><div class="card-title">Pedidos</div><div class="card-desc">Pedidos recebidos · status · valores · compradores · 30 dias.</div><span class="card-tag">/orders/search</span></div>
  <div class="card" onclick="abrirPerguntas()"><span class="card-ico">❓</span><div class="card-title">Perguntas</div><div class="card-desc">Sem resposta — responda direto daqui em 1 clique.</div><span class="card-tag">/questions/search</span></div>
  <div class="card" onclick="abrirMeusAnuncios()"><span class="card-ico">📋</span><div class="card-title">Meus Anúncios</div><div class="card-desc">Lista completa · pausar · ativar · editar estoque.</div><span class="card-tag">/users/items/search</span></div>
  <div class="card" onclick="abrirReputacao()"><span class="card-ico">⭐</span><div class="card-title">Reputação</div><div class="card-desc">Nível · cancelamentos · atrasos · capacidade de anúncios.</div><span class="card-tag">/marketplace/cap</span></div>
  <div class="card" onclick="abrirFaturamento()"><span class="card-ico">💰</span><div class="card-title">Faturamento</div><div class="card-desc">Cobrado · créditos · saldo do período atual no ML.</div><span class="card-tag">/billing</span></div>
  <div class="card" onclick="abrirMetricas()"><span class="card-ico">📊</span><div class="card-title">Métricas de Visitas</div><div class="card-desc">Visitas · conversão · top · sem visitas · baixa conversão.</div><span class="card-tag">/items_visits</span></div>
  <div class="card" onclick="abrirBenchmark()"><span class="card-ico">💵</span><div class="card-title">Price to Win</div><div class="card-desc">Preço competitivo oficial ML · identifica caros e baratos.</div><span class="card-tag">/price_to_win</span></div>
  <div class="card" onclick="abrirWebhooks()"><span class="card-ico">🔔</span><div class="card-title">Eventos Webhook</div><div class="card-desc">Eventos em tempo real do ML · orders · items · questions.</div><span class="card-tag">/webhook/ml</span></div>
  <div class="card" onclick="abrirCriarAnuncio()"><span class="card-ico">📝</span><div class="card-title">Publicar Anúncio</div><div class="card-desc">Cria anúncio direto no ML · categoria sugerida · validação.</div><span class="card-tag">POST /items</span></div>
  <div class="card" style="border-color:#f59e0b" onclick="abrirEspiao()"><span class="card-ico">🕵️</span><div class="card-title">Espião de Anúncios</div><div class="card-desc">Raio-X de qualquer MLB: vendas totais, preço, vendedor, reputação. Vigie concorrentes.</div><span class="card-tag">mini-Metrify</span></div>
  <div class="card" style="border-color:#8b5cf6" onclick="abrirWatchlist()"><span class="card-ico">👁️</span><div class="card-title">Watchlist</div><div class="card-desc">Concorrentes monitorados · vendas últimos 7/30 dias · variação de preço.</div><span class="card-tag">snapshot diário</span></div>
  <div class="card" style="border-color:#ec4899" onclick="abrirRanking()"><span class="card-ico">🏆</span><div class="card-title">Top Vendas</div><div class="card-desc">Ranking de mais vendidos por busca ou categoria · receita estimada.</div><span class="card-tag">/sites/MLB/search</span></div>
</div>

<script>
const fmt = (v,d=2) => v!=null ? parseFloat(v).toFixed(d).replace('.',',') : '-';

function showToast(msg, erro=false){
  document.querySelectorAll('.toast').forEach(t=>t.remove());
  const t = document.createElement('div');
  t.className='toast'; t.textContent=msg;
  t.style.background = erro ? '#450a0a' : '#064e3b';
  t.style.color = erro ? '#f87171' : '#4ade80';
  t.style.border = `1px solid ${erro?'#f87171':'#4ade80'}`;
  document.body.appendChild(t);
  setTimeout(()=>t.remove(), 3500);
}

// Status de conexão ML no hero
fetch('/api/ml-status').then(r=>r.json()).then(d=>{
  const el = document.getElementById('ml-status');
  if(!el) return;
  if(d.conectado){ el.textContent = '✅ Conectado: ' + (d.apelido || d.user_id || 'OK'); el.style.color = '#4ade80'; }
  else { el.textContent = '❌ Desconectado — acesse /config'; el.style.color = '#f87171'; }
}).catch(()=>{});

// ── Pedidos ───────────────────────────────────────────────────────
function abrirPedidos(dias=30, statusFiltro=''){
  showToast('🛒 Carregando pedidos...');
  fetch(`/api/pedidos?dias=${dias}&status=${statusFiltro}`)
  .then(r=>r.json()).then(d=>{ if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalPedidos(d); })
  .catch(()=>showToast('❌ Erro',true));
}

function mostrarModalPedidos(d){
  const fmt2=v=>v!=null?parseFloat(v).toFixed(2).replace('.',','):'0,00';
  const rows=(d.pedidos||[]).map(p=>`
    <tr style="border-bottom:1px solid #1a1f3a">
      <td style="padding:4px 7px;color:${p.status_cor};font-size:.8rem">${p.status_icone} ${p.status_label}</td>
      <td style="padding:4px 7px;color:#8b92a5;font-size:.72rem;white-space:nowrap">${p.data}</td>
      <td style="padding:4px 7px;color:#e4e6eb;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${p.titulo}">${p.titulo}</td>
      <td style="padding:4px 7px;color:#8b92a5;font-size:.75rem">${p.comprador}</td>
      <td style="padding:4px 7px;color:#fbbf24;text-align:right">x${p.qtd}</td>
      <td style="padding:4px 7px;color:#4ade80;text-align:right;font-weight:700">R$ ${fmt2(p.valor_pago)}</td>
    </tr>`).join('');

  const statusOpts=['','paid','shipped','delivered','cancelled'].map(s=>{
    const lab={'':`Todos (${d.total_ml})`,paid:'Pago',shipped:'Enviado',delivered:'Entregue',cancelled:'Cancelado'}[s]||s;
    return `<option value="${s}">${lab}</option>`;
  }).join('');

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #1d4ed8;width:900px;max-height:85vh" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div><h3 style="color:#60a5fa">🛒 Pedidos — últimos ${d.periodo_dias} dias</h3>
      <div style="color:#8b92a5;font-size:.72rem">${d.total_ml} pedidos no total · ${d.gerado_em}</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Total Período</div><div style="font-size:1.3rem;font-weight:700;color:#4ade80">R$ ${fmt2(d.total_valor)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Pedidos Hoje</div><div style="font-size:1.3rem;font-weight:700;color:#60a5fa">${d.qtd_hoje}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Semana (R$)</div><div style="font-size:1.3rem;font-weight:700;color:#fbbf24">R$ ${fmt2(d.val_semana)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Carregados</div><div style="font-size:1.3rem;font-weight:700;color:#c084fc">${d.total_processados}</div></div>
    </div>
    <div style="display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap">
      <select id="ped-dias" style="background:#1a1f3a;border:1px solid #2d3452;color:#e4e6eb;padding:4px 8px;border-radius:4px;font-size:.78rem">
        <option value="7">7 dias</option><option value="30" selected>30 dias</option><option value="60">60 dias</option><option value="90">90 dias</option>
      </select>
      <select id="ped-status" style="background:#1a1f3a;border:1px solid #2d3452;color:#e4e6eb;padding:4px 8px;border-radius:4px;font-size:.78rem">${statusOpts}</select>
      <button class="btn btn-blue" style="padding:4px 12px" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirPedidos(document.getElementById('ped-dias').value,document.getElementById('ped-status').value)">Filtrar</button>
    </div>
    <div style="overflow-y:auto;max-height:400px">
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.8rem">
      <thead><tr style="border-bottom:1px solid #2d3452;position:sticky;top:0;background:#0a0e27">
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Status</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Data</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Produto</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Comprador</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Qtd</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Valor</th>
      </tr></thead>
      <tbody>${rows||'<tr><td colspan="6" style="padding:20px;text-align:center;color:#8b92a5">Nenhum pedido encontrado.</td></tr>'}</tbody>
    </table></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

// ── Perguntas ─────────────────────────────────────────────────────
function abrirPerguntas(todas=false){
  showToast('❓ Carregando perguntas...');
  fetch(`/api/perguntas?todas=${todas?1:0}`)
  .then(r=>r.json()).then(d=>{ if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalPerguntas(d,todas); })
  .catch(()=>showToast('❌ Erro',true));
}

function mostrarModalPerguntas(d, todas=false){
  const rows=(d.perguntas||[]).map(p=>`
    <tr id="perg-row-${p.id}" style="border-bottom:1px solid #1a1f3a;${p.respondida?'opacity:.6':''}">
      <td style="padding:5px 7px;color:#8b92a5;font-size:.7rem;white-space:nowrap">${p.data}</td>
      <td style="padding:5px 7px;color:#e4e6eb;max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:.72rem" title="${p.item_titulo}">${p.item_titulo||p.item_id}</td>
      <td style="padding:5px 7px;color:#e4e6eb;font-size:.8rem">${p.texto}</td>
      <td style="padding:5px 7px;color:#8b92a5;font-size:.7rem">${p.from_user}</td>
      <td style="padding:5px 7px;min-width:240px">
        ${p.respondida
          ? `<span style="color:#4ade80;font-size:.72rem">✅ ${p.resposta_texto.substring(0,80)}${p.resposta_texto.length>80?'...':''}</span>`
          : `<div style="display:flex;gap:4px">
               <input id="resp-${p.id}" placeholder="Digite sua resposta..." style="flex:1;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:4px 6px;border-radius:4px;font-size:.75rem">
               <button onclick="enviarResposta(${p.id})" style="background:#059669;border:none;color:#fff;padding:4px 8px;border-radius:4px;cursor:pointer;font-size:.72rem;white-space:nowrap">Enviar</button>
             </div>`}
      </td>
    </tr>`).join('');

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #b45309;width:1000px;max-height:85vh" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div><h3 style="color:#fbbf24">❓ Perguntas ${todas?'(Todas)':'Sem Resposta'}</h3>
      <div style="color:#8b92a5;font-size:.72rem">${d.total} total · <span style="color:${d.nao_respondidas>0?'#f87171':'#4ade80'}">${d.nao_respondidas} sem resposta</span> · ${d.gerado_em}</div></div>
      <div style="display:flex;gap:6px">
        <button class="btn btn-gray" style="font-size:.72rem" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirPerguntas(${!todas})">${todas?'Ver Pendentes':'Ver Todas'}</button>
        <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
      </div>
    </div>
    <div style="overflow-y:auto;max-height:520px">
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.8rem">
      <thead><tr style="border-bottom:1px solid #2d3452;position:sticky;top:0;background:#0a0e27">
        <th style="padding:5px 7px;color:#8b92a5">Data</th>
        <th style="padding:5px 7px;color:#8b92a5">Anúncio</th>
        <th style="padding:5px 7px;color:#8b92a5">Pergunta</th>
        <th style="padding:5px 7px;color:#8b92a5">De</th>
        <th style="padding:5px 7px;color:#8b92a5">Resposta</th>
      </tr></thead>
      <tbody>${rows||'<tr><td colspan="5" style="padding:20px;text-align:center;color:#4ade80">✅ Nenhuma pergunta pendente!</td></tr>'}</tbody>
    </table></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function enviarResposta(qid){
  const inp=document.getElementById('resp-'+qid);
  if(!inp||!inp.value.trim()){showToast('❌ Digite uma resposta',true);return;}
  const texto=inp.value.trim();
  inp.disabled=true;
  fetch('/api/responder-pergunta',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({question_id:qid,resposta:texto})})
  .then(r=>r.json()).then(d=>{
    if(d.ok){
      const row=document.getElementById('perg-row-'+qid);
      if(row){
        row.style.opacity='.6';
        const td=row.querySelector('td:last-child');
        if(td) td.innerHTML=`<span style="color:#4ade80;font-size:.72rem">✅ ${texto.substring(0,80)}</span>`;
      }
      showToast('✅ Resposta enviada!');
    } else {
      inp.disabled=false;
      showToast('❌ '+d.erro,true);
    }
  }).catch(()=>{inp.disabled=false;showToast('❌ Erro',true);});
}

// ── Meus Anúncios ─────────────────────────────────────────────────
function abrirMeusAnuncios(status='active', offset=0){
  showToast('📋 Carregando anúncios...');
  fetch(`/api/meus-anuncios?status=${status}&offset=${offset}`)
  .then(r=>r.json()).then(d=>{ if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalAnuncios(d,status,offset); })
  .catch(()=>showToast('❌ Erro',true));
}

function mostrarModalAnuncios(d, statusAtual, offset){
  const fmt2=v=>v!=null?parseFloat(v).toFixed(2).replace('.',','):'0,00';
  const rows=(d.anuncios||[]).map(a=>`
    <tr style="border-bottom:1px solid #1a1f3a">
      <td style="padding:4px 6px;width:32px">${a.thumbnail?`<img src="${a.thumbnail}" style="width:28px;height:28px;object-fit:cover;border-radius:3px">`:'📦'}</td>
      <td style="padding:4px 7px;color:#e4e6eb;max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:.78rem" title="${a.titulo}">${a.titulo}</td>
      <td style="padding:4px 7px;color:#4ade80;text-align:right;font-weight:700">R$ ${fmt2(a.preco)}</td>
      <td style="padding:4px 7px;text-align:center"><span style="color:${a.estoque>5?'#4ade80':a.estoque>0?'#fbbf24':'#f87171'};font-weight:700">${a.estoque}</span></td>
      <td style="padding:4px 7px;color:#fbbf24;text-align:center">${a.vendas}</td>
      <td style="padding:4px 7px;font-size:.72rem">${a.tipo_icone} ${a.tipo_label}</td>
      <td style="padding:4px 7px;color:${a.status_cor};font-size:.72rem">${a.status==='active'?'🟢 Ativo':a.status==='paused'?'🟡 Pausado':'🔴 '+a.status}</td>
      <td style="padding:4px 4px;white-space:nowrap">
        ${a.status==='active'
          ? `<button onclick="acaoAnuncio('pausar','${a.id}',this)" style="background:#78350f;border:none;color:#fbbf24;padding:2px 7px;border-radius:3px;cursor:pointer;font-size:.7rem">⏸ Pausar</button>`
          : `<button onclick="acaoAnuncio('ativar','${a.id}',this)" style="background:#064e3b;border:none;color:#4ade80;padding:2px 7px;border-radius:3px;cursor:pointer;font-size:.7rem">▶ Ativar</button>`}
        <a href="${a.link}" target="_blank" style="color:#667eea;font-size:.72rem;margin-left:4px">↗</a>
      </td>
    </tr>`).join('');

  const total=d.total||0;
  const pagAtual=Math.floor(offset/50)+1;
  const totalPags=Math.ceil(total/50);
  const navPag=totalPags>1?`
    <div style="display:flex;gap:6px;justify-content:center;margin-top:8px">
      ${offset>0?`<button class="btn btn-gray" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirMeusAnuncios('${statusAtual}',${offset-50})">← Anterior</button>`:''}
      <span style="color:#8b92a5;padding:4px 8px;font-size:.78rem">${pagAtual}/${totalPags}</span>
      ${offset+50<total?`<button class="btn btn-gray" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirMeusAnuncios('${statusAtual}',${offset+50})">Próxima →</button>`:''}
    </div>`:'';

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #134e4a;width:950px;max-height:85vh" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div><h3 style="color:#34d399">📋 Meus Anúncios ML</h3>
      <div style="color:#8b92a5;font-size:.72rem">${total} anúncios · mostrando ${d.retornados} · ${d.gerado_em}</div></div>
      <div style="display:flex;gap:6px">
        <select id="an-status" style="background:#1a1f3a;border:1px solid #2d3452;color:#e4e6eb;padding:4px 8px;border-radius:4px;font-size:.78rem">
          <option value="active" ${statusAtual==='active'?'selected':''}>Ativos</option>
          <option value="paused" ${statusAtual==='paused'?'selected':''}>Pausados</option>
          <option value="closed" ${statusAtual==='closed'?'selected':''}>Fechados</option>
        </select>
        <button class="btn btn-cyan" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirMeusAnuncios(document.getElementById('an-status').value,0)">Filtrar</button>
        <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
      </div>
    </div>
    <div style="overflow-y:auto;max-height:500px">
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.8rem">
      <thead><tr style="border-bottom:1px solid #2d3452;position:sticky;top:0;background:#0a0e27">
        <th style="padding:5px 6px;color:#8b92a5;width:32px"></th>
        <th style="padding:5px 7px;color:#8b92a5">Título</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Preço</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:center">Estoque</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:center">Vendas</th>
        <th style="padding:5px 7px;color:#8b92a5">Tipo</th>
        <th style="padding:5px 7px;color:#8b92a5">Status</th>
        <th style="padding:5px 7px;color:#8b92a5">Ações</th>
      </tr></thead>
      <tbody>${rows||'<tr><td colspan="8" style="padding:20px;text-align:center;color:#8b92a5">Nenhum anúncio encontrado.</td></tr>'}</tbody>
    </table></div>
    ${navPag}
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function acaoAnuncio(acao, itemId, btn){
  btn.disabled=true; btn.textContent='⏳';
  const url=acao==='pausar'?'/api/pausar-anuncio':'/api/ativar-anuncio';
  fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({item_id:itemId})})
  .then(r=>r.json()).then(d=>{
    if(d.ok){ showToast('✅ '+d.msg); setTimeout(()=>abrirMeusAnuncios(),600); }
    else { btn.disabled=false; btn.textContent=acao==='pausar'?'⏸ Pausar':'▶ Ativar'; showToast('❌ '+d.erro,true); }
  }).catch(()=>{btn.disabled=false;showToast('❌ Erro',true);});
}

// ── Reputação ─────────────────────────────────────────────────────
function abrirReputacao(){
  showToast('⭐ Carregando reputação...');
  fetch('/api/reputacao').then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalReputacao(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function mostrarModalReputacao(d){
  const fmt1=v=>v!=null?parseFloat(v).toFixed(1).replace('.',','):'0,0';
  const metricaBloco=(label,valor,max,cor)=>{
    const pct=Math.min(100,(valor/max)*100);
    const bcor=valor>=(max*0.7)?'#f87171':valor>=(max*0.4)?'#fbbf24':'#4ade80';
    return `<div style="background:#0a0e27;padding:10px;border-radius:6px">
      <div style="display:flex;justify-content:space-between;margin-bottom:4px">
        <span style="font-size:.72rem;color:#8b92a5">${label}</span>
        <span style="font-size:.85rem;font-weight:700;color:${bcor}">${fmt1(valor)}%</span>
      </div>
      <div style="background:#2d3452;border-radius:3px;height:5px">
        <div style="background:${bcor};height:5px;border-radius:3px;width:${pct}%"></div>
      </div>
    </div>`;
  };
  const alertasH=(d.alertas||[]).map(a=>`
    <div style="background:${a.tipo==='critico'?'#450a0a':a.tipo==='aviso'?'#3b2700':'#064e3b'};border-radius:5px;padding:8px 12px;display:flex;align-items:center;gap:8px">
      <span>${a.icone}</span><span style="color:${a.cor};font-size:.82rem">${a.msg}</span>
    </div>`).join('');

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #7c3aed;width:680px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div>
        <h3 style="color:#c084fc">⭐ Reputação do Vendedor</h3>
        <div style="color:#8b92a5;font-size:.72rem">${d.apelido} ${d.power_icone} ${d.power_label} · ${d.gerado_em}</div>
      </div>
      <div style="text-align:center;padding:8px 16px;background:#0a0e27;border-radius:8px">
        <div style="font-size:2rem">${d.nivel_icone}</div>
        <div style="font-size:.82rem;font-weight:700;color:${d.nivel_cor}">${d.nivel_label}</div>
      </div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="display:flex;flex-direction:column;gap:6px;margin-bottom:12px">${alertasH}</div>
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Transações</div><div style="font-size:1.4rem;font-weight:800;color:#667eea">${d.total_transacoes||0}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Completas</div><div style="font-size:1.4rem;font-weight:800;color:#4ade80">${d.transacoes_completas||0}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Avaliação +</div><div style="font-size:1.4rem;font-weight:800;color:#fbbf24">${d.pct_positivo||0}%</div></div>
    </div>
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px">
      ${metricaBloco('Cancelamentos',d.pct_cancelamentos,5,'#f87171')}
      ${metricaBloco('Atrasos Envio',d.pct_atrasos,15,'#fbbf24')}
      ${metricaBloco('Reclamações',d.pct_reclamacoes,5,'#f87171')}
    </div>
    ${d.quota_anuncios?`<div style="background:#0a0e27;padding:10px;border-radius:6px;display:flex;justify-content:space-between;align-items:center">
      <span style="font-size:.78rem;color:#8b92a5">Capacidade de Anúncios</span>
      <span style="font-size:.85rem"><span style="color:#4ade80;font-weight:700">${d.anuncios_ativos}</span><span style="color:#8b92a5"> / ${d.quota_anuncios} (${d.slots_livres} livres)</span></span>
    </div>`:''}
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

// ── Faturamento ───────────────────────────────────────────────────
function abrirFaturamento(){
  showToast('💰 Carregando faturamento...');
  fetch('/api/faturamento').then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalFaturamento(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function mostrarModalFaturamento(d){
  const fmt2=v=>v!=null?parseFloat(v).toFixed(2).replace('.',','):'0,00';
  const linhasH=(d.linhas||[]).map(l=>`
    <tr style="border-bottom:1px solid #1a1f3a">
      <td style="padding:5px 10px;color:#e4e6eb;font-size:.8rem">${l.nome}</td>
      <td style="padding:5px 10px;color:${l.cor};font-weight:700;text-align:right;font-size:.85rem">${l.valor>=0?'+':''} R$ ${fmt2(Math.abs(l.valor))}</td>
    </tr>`).join('');

  const corSaldo=d.saldo>=0?'#4ade80':'#f87171';
  const nota=d.nota?`<div style="background:#1e3a5f;border-radius:5px;padding:8px 12px;font-size:.75rem;color:#93c5fd;margin-bottom:10px">ℹ️ ${d.nota}</div>`:'';

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #78350f;width:600px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div><h3 style="color:#fbbf24">💰 Faturamento ML</h3>
      <div style="color:#8b92a5;font-size:.72rem">Período: ${d.periodo_key} · ${d.gerado_em}</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    ${nota}
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#450a0a;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Cobrado</div><div style="font-size:1.2rem;font-weight:700;color:#f87171">R$ ${fmt2(d.total_cobrado)}</div></div>
      <div style="background:#064e3b;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Créditos</div><div style="font-size:1.2rem;font-weight:700;color:#4ade80">R$ ${fmt2(d.total_credito)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center;border:1px solid ${corSaldo}40"><div style="font-size:.6rem;color:#8b92a5">Saldo</div><div style="font-size:1.2rem;font-weight:700;color:${corSaldo}">R$ ${fmt2(d.saldo)}</div></div>
    </div>
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;border-radius:6px;overflow:hidden">
      <thead><tr style="border-bottom:1px solid #2d3452">
        <th style="padding:6px 10px;color:#8b92a5;text-align:left;font-size:.75rem">Descrição</th>
        <th style="padding:6px 10px;color:#8b92a5;text-align:right;font-size:.75rem">Valor</th>
      </tr></thead>
      <tbody>${linhasH||'<tr><td colspan="2" style="padding:20px;text-align:center;color:#8b92a5">Sem dados de faturamento.</td></tr>'}</tbody>
    </table>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

// ── Métricas (Visitas e Conversão) ────────────────────────────────
function abrirMetricas(dias=30){
  showToast('📊 Carregando métricas...');
  fetch(`/api/metricas?dias=${dias}`).then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalMetricas(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function mostrarModalMetricas(d){
  const fmt1=v=>v!=null?parseFloat(v).toFixed(1).replace('.',','):'0,0';
  const fmt2=v=>v!=null?parseFloat(v).toFixed(2).replace('.',','):'0,00';
  const rowAn=(a)=>{
    const rpv=a.rs_por_visita||0;
    const corRpv=rpv>=10?'#4ade80':rpv>=5?'#fbbf24':'#f87171';
    const badge=a.atinge_benchmark?'<span style="background:#065f46;color:#4ade80;padding:1px 5px;border-radius:3px;font-size:.65rem;margin-left:3px">✓</span>':'';
    return `<tr style="border-bottom:1px solid #1a1f3a">
    <td style="padding:3px 5px">${a.thumbnail?`<img src="${a.thumbnail}" style="width:24px;height:24px;object-fit:cover;border-radius:3px">`:''}</td>
    <td style="padding:3px 7px;color:#e4e6eb;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:.75rem" title="${a.titulo}">${a.titulo}</td>
    <td style="padding:3px 7px;color:#60a5fa;text-align:right;font-weight:700">${a.visitas}</td>
    <td style="padding:3px 7px;color:#fbbf24;text-align:right">${a.vendas_total}</td>
    <td style="padding:3px 7px;color:${a.conversao_pct>=3?'#4ade80':a.conversao_pct>=1?'#fbbf24':'#f87171'};text-align:right;font-weight:700">${fmt1(a.conversao_pct)}%</td>
    <td style="padding:3px 7px;color:#4ade80;text-align:right">R$ ${fmt2(a.preco)}</td>
    <td style="padding:3px 7px;color:${corRpv};text-align:right;font-weight:700" title="Meta: R$ 10,00/visita">R$ ${fmt2(rpv)}${badge}</td>
    <td><a href="${a.link}" target="_blank" style="color:#667eea;font-size:.72rem">↗</a></td>
  </tr>`;};

  const renderBlocoLista=(titulo,lista,cor,emoji,vazio)=>{
    if(!lista||!lista.length)return`<div style="color:#8b92a5;text-align:center;padding:10px;font-size:.78rem">${vazio}</div>`;
    return `<div style="margin-bottom:14px"><div style="color:${cor};font-weight:700;font-size:.8rem;margin-bottom:6px">${emoji} ${titulo} (${lista.length})</div>
      <div style="max-height:180px;overflow-y:auto"><table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.75rem">
      <thead><tr style="border-bottom:1px solid #2d3452"><th style="width:30px"></th><th style="padding:3px 5px;color:#8b92a5;text-align:left">Anúncio</th><th style="padding:3px 5px;color:#8b92a5;text-align:right">Visitas</th><th style="padding:3px 5px;color:#8b92a5;text-align:right">Vendas</th><th style="padding:3px 5px;color:#8b92a5;text-align:right">Conv.</th><th style="padding:3px 5px;color:#8b92a5;text-align:right">Preço</th><th style="padding:3px 5px;color:#8b92a5;text-align:right" title="Receita R$ por visita (meta: R$ 10)">R$/Visita</th><th></th></tr></thead>
      <tbody>${lista.map(rowAn).join('')}</tbody></table></div></div>`;
  };

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #166534;width:920px;max-height:90vh" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div><h3 style="color:#4ade80">📊 Métricas de Visitas e Conversão</h3>
      <div style="color:#8b92a5;font-size:.72rem">Últimos ${d.periodo_dias} dias · ${d.gerado_em}</div></div>
      <div style="display:flex;gap:6px">
        <select id="met-dias" style="background:#1a1f3a;border:1px solid #2d3452;color:#e4e6eb;padding:4px 8px;border-radius:4px;font-size:.78rem">
          <option value="7">7 dias</option><option value="15">15 dias</option><option value="30" selected>30 dias</option><option value="60">60 dias</option>
        </select>
        <button class="btn btn-green" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirMetricas(document.getElementById('met-dias').value)">Atualizar</button>
        <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
      </div>
    </div>
    ${(()=>{
      const rpv=d.rs_por_visita||0;
      const meta=d.benchmark_rs_visita||10;
      const pct=d.pct_benchmark||0;
      const ok=d.atinge_benchmark_geral;
      const cor=ok?'#4ade80':pct>=70?'#fbbf24':'#f87171';
      const bg=ok?'#064e3b':pct>=70?'#3b2700':'#450a0a';
      const msg=ok?`✅ Acima da meta (R$ ${meta.toFixed(2).replace('.',',')}/visita)`:
                   `⚠️ Abaixo da meta — está em ${pct.toFixed(1).replace('.',',')}% de R$ ${meta.toFixed(2).replace('.',',')}/visita`;
      return `<div style="background:${bg};border:1px solid ${cor};padding:8px 12px;border-radius:6px;margin-bottom:10px;display:flex;justify-content:space-between;align-items:center">
        <div><div style="color:${cor};font-weight:700;font-size:.85rem">🎯 Benchmark: 1000 visitas = R$ 10.000 em vendas</div>
        <div style="color:#e4e6eb;font-size:.75rem">${msg}</div></div>
        <div style="text-align:right"><div style="color:#8b92a5;font-size:.65rem">Receita estimada (${d.periodo_dias}d)</div>
        <div style="color:${cor};font-weight:700;font-size:1rem">R$ ${(d.receita_estimada||0).toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})}</div></div>
      </div>`;
    })()}
    <div style="display:grid;grid-template-columns:repeat(6,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Total Visitas</div><div style="font-size:1.3rem;font-weight:700;color:#60a5fa">${(d.total_visitas_vendedor||0).toLocaleString('pt-BR')}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Média/Dia</div><div style="font-size:1.3rem;font-weight:700;color:#c084fc">${fmt1(d.media_diaria)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Vendas</div><div style="font-size:1.3rem;font-weight:700;color:#4ade80">${d.total_vendas_periodo||0}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Conv. Média</div><div style="font-size:1.3rem;font-weight:700;color:#fbbf24">${fmt1(d.conversao_media_pct)}%</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center" title="Meta: R$ 10,00 por visita"><div style="font-size:.6rem;color:#8b92a5">R$/Visita</div><div style="font-size:1.3rem;font-weight:700;color:${(d.rs_por_visita||0)>=10?'#4ade80':(d.rs_por_visita||0)>=5?'#fbbf24':'#f87171'}">R$ ${fmt2(d.rs_por_visita)}</div></div>
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Analisados</div><div style="font-size:1.3rem;font-weight:700;color:#8b92a5">${d.total_anuncios_analisados||0}</div></div>
    </div>
    ${renderBlocoLista('Top Visitados',d.top_anuncios,'#4ade80','🏆','Sem dados.')}
    ${renderBlocoLista('Visitados sem vender (problema de preço/ficha)',d.baixa_conversao,'#fbbf24','⚠️','Nenhum anúncio com baixa conversão.')}
    ${renderBlocoLista('Sem Visitas — Revisar SEO/categoria',d.sem_visitas,'#f87171','👻','Todos os anúncios têm visitas!')}
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

// ── Benchmark / Price to Win ──────────────────────────────────────
function abrirBenchmark(){
  showToast('💵 Analisando preços vs concorrência... (demora ~15s)');
  fetch('/api/benchmark?limite=30').then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalBenchmark(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function mostrarModalBenchmark(d){
  const fmt2=v=>v!=null?parseFloat(v).toFixed(2).replace('.',','):'0,00';
  const rows=(d.analise||[]).map(a=>{
    const diffStr=(a.diff_pct>0?'+':'')+a.diff_pct.toFixed(1).replace('.',',')+'%';
    return `<tr style="border-bottom:1px solid #1a1f3a">
      <td style="padding:4px 7px;font-size:1rem;text-align:center">${a.icone}</td>
      <td style="padding:4px 7px;color:#e4e6eb;max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:.78rem" title="${a.titulo}">${a.titulo}</td>
      <td style="padding:4px 7px;color:#e4e6eb;text-align:right">R$ ${fmt2(a.preco_atual)}</td>
      <td style="padding:4px 7px;color:#4ade80;text-align:right;font-weight:700">R$ ${fmt2(a.preco_sugerido)}</td>
      <td style="padding:4px 7px;color:${a.cor};text-align:right;font-weight:700">${diffStr}</td>
      <td style="padding:4px 7px;color:${a.cor};font-size:.72rem">${a.acao}</td>
      <td style="padding:4px 7px;color:#fbbf24;text-align:right">${a.vendas}</td>
      <td style="padding:4px 4px;white-space:nowrap">
        <button onclick="aplicarPrecoMl('${a.id}',${a.preco_sugerido},this)" style="background:#065f46;border:none;color:#4ade80;padding:2px 6px;border-radius:3px;cursor:pointer;font-size:.7rem">Aplicar</button>
        <a href="${a.link}" target="_blank" style="color:#667eea;font-size:.72rem;margin-left:4px">↗</a>
      </td>
    </tr>`;
  }).join('');

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #701a75;width:1000px;max-height:85vh" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div><h3 style="color:#e879f9">💵 Benchmark de Preços — Price to Win</h3>
      <div style="color:#8b92a5;font-size:.72rem">Preço oficial sugerido pelo ML para ganhar Buy Box · ${d.gerado_em}</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px">
      <div style="background:#0a0e27;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Analisados</div><div style="font-size:1.3rem;font-weight:700;color:#c084fc">${d.total_analisados||0}</div></div>
      <div style="background:#450a0a;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">🔴 Caros</div><div style="font-size:1.3rem;font-weight:700;color:#f87171">${d.caros||0}</div></div>
      <div style="background:#3b2700;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">🟡 Baratos</div><div style="font-size:1.3rem;font-weight:700;color:#fbbf24">${d.baratos||0}</div></div>
      <div style="background:#064e3b;padding:10px;border-radius:6px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">✅ OK</div><div style="font-size:1.3rem;font-weight:700;color:#4ade80">${d.competitivos||0}</div></div>
    </div>
    <div style="overflow-y:auto;max-height:500px">
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.8rem">
      <thead><tr style="border-bottom:1px solid #2d3452;position:sticky;top:0;background:#0a0e27">
        <th></th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Anúncio</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Seu Preço</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Preço Win</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Dif.</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Recomendação</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Vendas</th>
        <th></th>
      </tr></thead>
      <tbody>${rows||'<tr><td colspan="8" style="padding:20px;text-align:center;color:#8b92a5">Nenhum dado de benchmark disponível.</td></tr>'}</tbody>
    </table></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function aplicarPrecoMl(itemId, novoPreco, btn){
  if(!confirm(`Alterar preço do anúncio para R$ ${novoPreco.toFixed(2).replace('.',',')}?`)) return;
  btn.disabled=true; btn.textContent='⏳';
  fetch('/api/editar-anuncio',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({item_id:itemId,alteracoes:{preco:novoPreco}})})
  .then(r=>r.json()).then(d=>{
    if(d.ok){ btn.textContent='✅'; btn.style.background='#064e3b'; showToast('✅ Preço atualizado!'); }
    else { btn.disabled=false; btn.textContent='Aplicar'; showToast('❌ '+d.erro,true); }
  }).catch(()=>{btn.disabled=false;btn.textContent='Aplicar';showToast('❌ Erro',true);});
}

// ── Eventos Webhook ──────────────────────────────────────────────
function abrirWebhooks(topic=''){
  showToast('🔔 Carregando eventos...');
  fetch(`/api/webhooks-eventos?topic=${topic}`).then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;} mostrarModalWebhooks(d,topic);
  }).catch(()=>showToast('❌ Erro',true));
}

function mostrarModalWebhooks(d, topicAtual){
  const rows=(d.eventos||[]).map(e=>`<tr style="border-bottom:1px solid #1a1f3a">
    <td style="padding:4px 7px;color:${e.cor}">${e.icone} ${e.label}</td>
    <td style="padding:4px 7px;color:#8b92a5;font-size:.72rem;white-space:nowrap">${e.data}</td>
    <td style="padding:4px 7px;color:#e4e6eb;font-family:monospace;font-size:.72rem">${e.resource}</td>
    <td style="padding:4px 7px">${e.link?`<a href="${e.link}" target="_blank" style="color:#667eea;font-size:.75rem">↗</a>`:''}</td>
  </tr>`).join('');

  const statsKeys=Object.keys(d.stats||{});
  const filtros=`<button class="btn btn-gray" style="font-size:.7rem;padding:3px 7px" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirWebhooks('')">Todos</button>` +
    statsKeys.map(k=>{const cnt=d.stats[k];const ativo=k===topicAtual?'background:#991b1b':'';return `<button class="btn btn-gray" style="font-size:.7rem;padding:3px 7px;${ativo}" onclick="const bg=this.closest('.modal-bg');bg.remove();abrirWebhooks('${k}')">${k} (${cnt})</button>`}).join('');

  const webhookUrl = window.location.origin + '/webhook/ml';

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #991b1b;width:900px;max-height:88vh" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
      <div><h3 style="color:#fca5a5">🔔 Eventos em Tempo Real (Webhooks ML)</h3>
      <div style="color:#8b92a5;font-size:.72rem">Último evento: ${d.ultimo_evento||'nenhum'} · Total: ${d.total||0}</div></div>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <div style="background:#0a0e27;border:1px dashed #2d3452;border-radius:6px;padding:10px;margin-bottom:12px;font-size:.75rem">
      <div style="color:#fbbf24;font-weight:700;margin-bottom:4px">⚙️ Configure no painel ML:</div>
      <div style="color:#8b92a5">URL de Notificações:
        <input readonly value="${webhookUrl}" onclick="this.select()" style="background:#1a1f3a;border:1px solid #2d3452;color:#93c5fd;padding:3px 6px;border-radius:3px;width:60%;font-family:monospace;font-size:.72rem">
      </div>
      <div style="color:#8b92a5;margin-top:4px">Tópicos recomendados: <code style="color:#c084fc">orders_v2, items, questions, messages, shipments</code></div>
    </div>
    <div style="display:flex;gap:4px;margin-bottom:10px;flex-wrap:wrap">${filtros}</div>
    <div style="overflow-y:auto;max-height:440px">
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.8rem">
      <thead><tr style="border-bottom:1px solid #2d3452;position:sticky;top:0;background:#0a0e27">
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Tópico</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Recebido</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Recurso</th>
        <th></th>
      </tr></thead>
      <tbody>${rows||'<tr><td colspan="4" style="padding:30px;text-align:center;color:#8b92a5">Nenhum evento. Configure o webhook no painel ML e aguarde.</td></tr>'}</tbody>
    </table></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

// ── Criar Anúncio (publicar produto do catálogo no ML) ────────────
function abrirCriarAnuncio(){
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #0369a1;width:680px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
      <h3 style="color:#38bdf8">📝 Publicar Anúncio no ML</h3>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <p style="color:#8b92a5;font-size:.8rem;margin-bottom:12px">Publica um produto do seu catálogo diretamente no Mercado Livre. A categoria é sugerida automaticamente pelo título.</p>
    <label style="color:#8b92a5;font-size:.7rem;text-transform:uppercase">Título (máx 60 chars)</label>
    <input id="ca-titulo" maxlength="60" placeholder="Ex: Tênis Nike Air Max 90 Preto Original" style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:8px 10px;border-radius:5px;margin-bottom:10px;font-size:.85rem">
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:10px">
      <div><label style="color:#8b92a5;font-size:.7rem">Preço (R$)</label><input id="ca-preco" type="number" step="0.01" style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:8px 10px;border-radius:5px"></div>
      <div><label style="color:#8b92a5;font-size:.7rem">Estoque</label><input id="ca-qtd" type="number" value="1" min="1" style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:8px 10px;border-radius:5px"></div>
      <div><label style="color:#8b92a5;font-size:.7rem">Tipo</label>
        <select id="ca-tipo" style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:8px 10px;border-radius:5px">
          <option value="gold_special">Clássico</option><option value="gold_pro">Premium</option>
        </select></div>
    </div>
    <label style="color:#8b92a5;font-size:.7rem;text-transform:uppercase">Imagens (URLs, 1 por linha — mín 1, máx 12)</label>
    <textarea id="ca-imgs" rows="3" placeholder="https://exemplo.com/foto1.jpg" style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:8px 10px;border-radius:5px;margin-bottom:10px;font-size:.8rem;font-family:monospace"></textarea>
    <label style="color:#8b92a5;font-size:.7rem;text-transform:uppercase">Descrição</label>
    <textarea id="ca-desc" rows="4" placeholder="Descreva o produto..." style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:8px 10px;border-radius:5px;margin-bottom:10px;font-size:.8rem"></textarea>
    <div id="ca-cat-sugg" style="background:#0a0e27;border-radius:5px;padding:8px;margin-bottom:10px;font-size:.78rem;color:#8b92a5;display:none"></div>
    <div style="display:flex;gap:6px;justify-content:flex-end">
      <button class="btn btn-gray" onclick="preverCategoriaForm()">🔮 Sugerir Categoria</button>
      <button class="btn btn-yellow" onclick="validarAnuncioForm()">✓ Validar</button>
      <button class="btn btn-blue" onclick="publicarAnuncioForm()">📝 Publicar</button>
    </div>
    <div id="ca-result" style="margin-top:12px"></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

function _caDados(){
  return {
    titulo: document.getElementById('ca-titulo').value.trim(),
    preco: parseFloat(document.getElementById('ca-preco').value||0),
    quantidade: parseInt(document.getElementById('ca-qtd').value||1),
    tipo_anuncio: document.getElementById('ca-tipo').value,
    imagens: document.getElementById('ca-imgs').value.split('\n').map(s=>s.trim()).filter(Boolean),
    descricao: document.getElementById('ca-desc').value.trim(),
  };
}

function preverCategoriaForm(){
  const titulo=document.getElementById('ca-titulo').value.trim();
  if(titulo.length<4){showToast('❌ Título curto',true);return;}
  showToast('🔮 Buscando categoria...');
  fetch('/api/prever-categoria',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({titulo})})
  .then(r=>r.json()).then(d=>{
    const box=document.getElementById('ca-cat-sugg');
    if(!d.ok){box.style.display='block';box.innerHTML='❌ '+d.erro;return;}
    box.style.display='block';
    const sugs=(d.sugestoes||[]).map((s,i)=>`
      <div style="display:flex;justify-content:space-between;padding:3px 0;border-bottom:1px solid #1a1f3a">
        <span style="color:${i===0?'#4ade80':'#e4e6eb'}">${i===0?'⭐ ':''}${s.category_name} <span style="color:#8b92a5">(${s.category_id})</span></span>
        <span style="color:#8b92a5;font-size:.7rem">${s.domain_name}</span>
      </div>`).join('');
    box.innerHTML=`<div style="color:#4ade80;font-weight:700;margin-bottom:4px">✅ Categoria sugerida: <code>${d.top_category_id}</code> — ${d.top_category_name}</div>${sugs}`;
    window._caCategoriaId=d.top_category_id;
    showToast('✅ Categoria: '+d.top_category_name);
  }).catch(()=>showToast('❌ Erro',true));
}

function validarAnuncioForm(){
  if(!window._caCategoriaId){showToast('❌ Sugira a categoria antes',true);return;}
  const dados=_caDados(); dados.category_id=window._caCategoriaId;
  if(!dados.titulo||!dados.preco||!dados.imagens.length){showToast('❌ Título, preço e imagem obrigatórios',true);return;}
  showToast('✓ Validando...');
  const payload={title:dados.titulo,category_id:dados.category_id,price:dados.preco,currency_id:'BRL',available_quantity:dados.quantidade,buying_mode:'buy_it_now',listing_type_id:dados.tipo_anuncio,condition:'new',pictures:dados.imagens.map(u=>({source:u}))};
  fetch('/api/validar-anuncio',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)})
  .then(r=>r.json()).then(d=>{
    const res=document.getElementById('ca-result');
    if(d.ok){res.innerHTML=`<div style="background:#064e3b;color:#4ade80;padding:8px;border-radius:5px">✅ ${d.msg}</div>`;}
    else{res.innerHTML=`<div style="background:#450a0a;color:#f87171;padding:8px;border-radius:5px">❌ ${d.erro}</div>`;}
  }).catch(()=>showToast('❌ Erro',true));
}

function publicarAnuncioForm(){
  if(!window._caCategoriaId){showToast('❌ Sugira a categoria antes',true);return;}
  const dados=_caDados(); dados.category_id=window._caCategoriaId;
  if(!dados.titulo||!dados.preco||!dados.imagens.length){showToast('❌ Campos obrigatórios faltando',true);return;}
  if(!confirm(`Publicar "${dados.titulo}" por R$ ${dados.preco.toFixed(2).replace('.',',')}?`)) return;
  showToast('📝 Publicando...');
  fetch('/api/criar-anuncio',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(dados)})
  .then(r=>r.json()).then(d=>{
    const res=document.getElementById('ca-result');
    if(d.ok){res.innerHTML=`<div style="background:#064e3b;color:#4ade80;padding:10px;border-radius:5px">✅ ${d.msg} · <a href="${d.link}" target="_blank" style="color:#60a5fa">Ver anúncio →</a></div>`;showToast('✅ Publicado!');}
    else{res.innerHTML=`<div style="background:#450a0a;color:#f87171;padding:10px;border-radius:5px">❌ ${d.erro}</div>`;}
  }).catch(()=>showToast('❌ Erro',true));
}

// ══════════════════════════════════════════════════════════════════
// ESPIÃO — mini-Metrify
// ══════════════════════════════════════════════════════════════════
const BENCHMARK_RS_POR_VISITA = 10.0;  // meta: 1000 visitas = R$ 10k

function abrirEspiao(){
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #f59e0b;width:560px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
      <h3 style="color:#fbbf24">🕵️ Espião de Anúncios</h3>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <p style="color:#8b92a5;font-size:.8rem;margin-bottom:12px">Cole um <code style="color:#fbbf24">MLBxxxxxxxxxx</code>, link do anúncio ou URL de busca. Dá pra espiar qualquer vendedor.</p>
    <input id="spy-input" placeholder="MLB1234567890 ou https://produto.mercadolivre.com.br/..." style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:10px;border-radius:5px;margin-bottom:10px;font-family:monospace;font-size:.85rem">
    <div style="display:flex;gap:6px;justify-content:flex-end;margin-bottom:10px">
      <button class="btn btn-yellow" onclick="executarSpy()">🔍 Espionar</button>
    </div>
    <div id="spy-result"></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
  setTimeout(()=>document.getElementById('spy-input').focus(),100);
  document.getElementById('spy-input').addEventListener('keydown',e=>{if(e.key==='Enter')executarSpy();});
}

function executarSpy(){
  const v=document.getElementById('spy-input').value.trim();
  if(!v){showToast('❌ Informe um MLB ou link',true);return;}
  showToast('🕵️ Espionando...');
  fetch('/api/spy-anuncio',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({item_id:v})})
  .then(r=>r.json()).then(d=>{
    const res=document.getElementById('spy-result');
    if(!d.ok){res.innerHTML=`<div style="background:#450a0a;color:#f87171;padding:10px;border-radius:5px">❌ ${d.erro}</div>`;return;}
    res.innerHTML=renderSpy(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function renderSpy(d){
  const ved=d.vendedor||{};
  const delta=d.delta||{};
  const badgeWatch=d.monitorado
    ? `<button class="btn btn-gray" style="font-size:.7rem" onclick="toggleWatch('${d.item_id}',false)">🗑️ Remover da watchlist</button>`
    : `<button class="btn" style="background:#8b5cf6;color:#fff;font-size:.7rem" onclick="toggleWatch('${d.item_id}',true)">👁️ Monitorar diariamente</button>`;
  const statusCor=d.status==='active'?'#4ade80':d.status==='paused'?'#fbbf24':'#f87171';
  const rev7=delta.tem_historico?`<div style="background:#064e3b;padding:8px;border-radius:5px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Vendas 7d</div><div style="font-size:1.3rem;font-weight:700;color:#4ade80">${delta.vendas_7d||0}</div><div style="font-size:.7rem;color:#4ade80">R$ ${fmt(d.receita_7d||0)}</div></div>`
    :`<div style="background:#1c1917;padding:8px;border-radius:5px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Vendas 7d</div><div style="font-size:.85rem;color:#8b92a5">sem histórico<br>ainda</div></div>`;
  const priceDiff=delta.tem_historico && delta.preco_diff_7d!==0
    ? `<span style="color:${delta.preco_diff_7d>0?'#f87171':'#4ade80'};font-size:.7rem">${delta.preco_diff_7d>0?'▲':'▼'} R$ ${fmt(Math.abs(delta.preco_diff_7d))} em 7d</span>`
    : '';
  return `<div style="background:#0a0e27;border-radius:6px;padding:14px;margin-top:6px">
    <div style="display:flex;gap:10px;margin-bottom:10px">
      ${d.thumbnail?`<img src="${d.thumbnail}" style="width:70px;height:70px;object-fit:cover;border-radius:5px">`:''}
      <div style="flex:1">
        <div style="font-weight:700;color:#e4e6eb;font-size:.88rem;line-height:1.3">${d.titulo}</div>
        <div style="color:#8b92a5;font-size:.7rem;margin-top:2px"><code style="color:#fbbf24">${d.item_id}</code> · ${d.categoria_nome||d.category_id}</div>
        <div style="color:${statusCor};font-size:.72rem;margin-top:3px">${d.status==='active'?'🟢 Ativo':d.status==='paused'?'🟡 Pausado':'🔴 '+d.status} · ${d.condicao} · ${d.tipo_anuncio}</div>
      </div>
    </div>
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:6px;margin-bottom:10px">
      <div style="background:#1a1f3a;padding:8px;border-radius:5px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Preço</div><div style="font-size:1.1rem;font-weight:700;color:#4ade80">R$ ${fmt(d.preco)}</div>${priceDiff}</div>
      <div style="background:#1a1f3a;padding:8px;border-radius:5px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Vendas TOTAIS</div><div style="font-size:1.3rem;font-weight:800;color:#fbbf24">${d.vendas_total}</div><div style="font-size:.65rem;color:#8b92a5">acumulado</div></div>
      <div style="background:#1a1f3a;padding:8px;border-radius:5px;text-align:center"><div style="font-size:.6rem;color:#8b92a5">Estoque</div><div style="font-size:1.1rem;font-weight:700;color:${d.estoque>5?'#4ade80':d.estoque>0?'#fbbf24':'#f87171'}">${d.estoque}</div></div>
      ${rev7}
    </div>
    <div style="background:#1a1f3a;padding:10px;border-radius:5px;margin-bottom:10px">
      <div style="color:#8b92a5;font-size:.7rem;margin-bottom:4px">💰 Receita acumulada estimada</div>
      <div style="font-size:1.2rem;font-weight:800;color:#c084fc">R$ ${fmt(d.receita_acum)}</div>
      <div style="color:#8b92a5;font-size:.65rem">vendas totais × preço atual (estimativa)</div>
    </div>
    <div style="background:#1a1f3a;padding:10px;border-radius:5px;margin-bottom:10px">
      <div style="color:#8b92a5;font-size:.7rem;margin-bottom:4px">🏪 Vendedor</div>
      <div style="font-size:.88rem;color:#e4e6eb">${ved.apelido||'?'} · <span style="color:#8b92a5">${ved.nivel||'-'}</span></div>
      <div style="color:#8b92a5;font-size:.7rem">${(ved.transacoes||0).toLocaleString('pt-BR')} transações · ${ved.positivas_pct||0}% positivas ${ved.cidade?' · '+ved.cidade:''}</div>
    </div>
    <div style="display:flex;gap:6px;justify-content:space-between;align-items:center">
      <a href="${d.link}" target="_blank" class="btn btn-blue" style="font-size:.72rem">↗ Abrir no ML</a>
      ${badgeWatch}
    </div>
    ${!delta.tem_historico?'<div style="color:#8b92a5;font-size:.7rem;margin-top:8px;text-align:center">💡 Monitore para ver "vendas no período" a partir de amanhã</div>':''}
  </div>`;
}

function toggleWatch(itemId, add){
  const url=add?'/api/watch-add':'/api/watch-remove';
  fetch(url,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({item_id:itemId})})
  .then(r=>r.json()).then(d=>{
    if(d.ok){showToast(add?'👁️ Adicionado à watchlist':'🗑️ Removido');executarSpy();}
    else{showToast('❌ '+d.erro,true);}
  });
}

// ══════════════════════════════════════════════════════════════════
// WATCHLIST
// ══════════════════════════════════════════════════════════════════
function abrirWatchlist(){
  showToast('👁️ Atualizando watchlist... (snapshot de cada item)');
  fetch('/api/watchlist').then(r=>r.json()).then(d=>{
    if(!d.ok){showToast('❌ '+d.erro,true);return;}
    mostrarModalWatchlist(d);
  }).catch(()=>showToast('❌ Erro',true));
}

function mostrarModalWatchlist(d){
  const rows=(d.itens||[]).map(i=>{
    const priceDiff=i.preco_diff_7d && i.preco_diff_7d!==0
      ? `<span style="color:${i.preco_diff_7d>0?'#f87171':'#4ade80'};font-size:.66rem">${i.preco_diff_7d>0?'▲':'▼'}${fmt(Math.abs(i.preco_diff_7d))}</span>`
      : '';
    return `<tr style="border-bottom:1px solid #1a1f3a">
      <td style="padding:4px 6px">${i.thumbnail?`<img src="${i.thumbnail}" style="width:32px;height:32px;object-fit:cover;border-radius:3px">`:''}</td>
      <td style="padding:4px 7px;max-width:180px">
        <div style="color:#e4e6eb;font-size:.76rem;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${i.titulo}">${i.apelido||i.titulo}</div>
        <div style="color:#8b92a5;font-size:.64rem"><code>${i.item_id}</code></div>
      </td>
      <td style="padding:4px 7px;color:#4ade80;text-align:right;font-weight:700">R$ ${fmt(i.preco)}<br>${priceDiff}</td>
      <td style="padding:4px 7px;color:#fbbf24;text-align:right;font-weight:700">${i.sold_total}</td>
      <td style="padding:4px 7px;color:${i.vendas_7d>0?'#4ade80':'#8b92a5'};text-align:right;font-weight:700">${i.tem_historico?'+'+i.vendas_7d:'—'}<br><span style="font-size:.62rem;color:#8b92a5">${i.tem_historico?'R$ '+fmt(i.receita_7d):'aguarde'}</span></td>
      <td style="padding:4px 7px;color:#c084fc;text-align:right">${i.tem_historico?'+'+i.vendas_30d:'—'}</td>
      <td style="padding:4px 7px;color:${i.estoque>5?'#4ade80':i.estoque>0?'#fbbf24':'#f87171'};text-align:center">${i.estoque}</td>
      <td style="padding:4px 4px;white-space:nowrap">
        <a href="${i.link}" target="_blank" style="color:#667eea;font-size:.72rem;margin-right:4px">↗</a>
        <button onclick="toggleWatch('${i.item_id}',false);setTimeout(abrirWatchlist,500)" style="background:#450a0a;border:none;color:#f87171;padding:2px 6px;border-radius:3px;cursor:pointer;font-size:.65rem">🗑️</button>
      </td>
    </tr>`;
  }).join('');

  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #8b5cf6;width:1000px;max-height:88vh" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
      <div><h3 style="color:#c084fc">👁️ Watchlist de Concorrentes</h3>
      <div style="color:#8b92a5;font-size:.72rem">${d.total} anúncios monitorados · snapshot diário · ${d.gerado_em}</div></div>
      <div style="display:flex;gap:6px">
        <button class="btn" style="background:#8b5cf6;color:#fff" onclick="abrirEspiao()">+ Adicionar</button>
        <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
      </div>
    </div>
    <div style="overflow-y:auto;max-height:560px">
    <table style="width:100%;border-collapse:collapse;background:#0a0e27;font-size:.8rem">
      <thead><tr style="border-bottom:1px solid #2d3452;position:sticky;top:0;background:#0a0e27">
        <th style="width:40px"></th>
        <th style="padding:6px 7px;color:#8b92a5;text-align:left">Anúncio</th>
        <th style="padding:6px 7px;color:#8b92a5;text-align:right">Preço</th>
        <th style="padding:6px 7px;color:#8b92a5;text-align:right">Vendas Tot</th>
        <th style="padding:6px 7px;color:#8b92a5;text-align:right">7d</th>
        <th style="padding:6px 7px;color:#8b92a5;text-align:right">30d</th>
        <th style="padding:6px 7px;color:#8b92a5;text-align:center">Estoque</th>
        <th></th>
      </tr></thead>
      <tbody>${rows||'<tr><td colspan="8" style="padding:30px;text-align:center;color:#8b92a5">Watchlist vazia. Use o <b>Espião</b> para adicionar anúncios.</td></tr>'}</tbody>
    </table></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
}

// ══════════════════════════════════════════════════════════════════
// RANKING — Top Vendas por Busca/Categoria
// ══════════════════════════════════════════════════════════════════
function abrirRanking(){
  const h=`<div class="modal-bg" onclick="this.remove()"><div class="modal" style="border:1px solid #ec4899;width:560px" onclick="event.stopPropagation()">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
      <h3 style="color:#f472b6">🏆 Top Vendas — Ranking</h3>
      <button class="btn btn-gray" onclick="this.closest('.modal-bg').remove()">✕</button>
    </div>
    <p style="color:#8b92a5;font-size:.8rem;margin-bottom:12px">Descubra quem mais vende num nicho. Por busca (ex: <i>tênis nike</i>) ou categoria (ex: <code>MLB1276</code>).</p>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:10px">
      <button class="btn btn-gray" onclick="document.getElementById('rk-tipo').value='busca';document.getElementById('rk-input').placeholder='tênis nike air max'">🔎 Por Busca</button>
      <button class="btn btn-gray" onclick="document.getElementById('rk-tipo').value='categoria';document.getElementById('rk-input').placeholder='MLB1276'">📁 Por Categoria</button>
    </div>
    <input type="hidden" id="rk-tipo" value="busca">
    <input id="rk-input" placeholder="tênis nike air max" style="width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:10px;border-radius:5px;margin-bottom:10px;font-size:.85rem">
    <div style="display:flex;gap:6px;justify-content:flex-end;margin-bottom:10px">
      <button class="btn" style="background:#ec4899;color:#fff" onclick="executarRanking()">🏆 Ranquear</button>
    </div>
    <div id="rk-result"></div>
  </div></div>`;
  document.querySelectorAll('.modal-bg').forEach(m=>m.remove());
  document.body.insertAdjacentHTML('beforeend',h);
  setTimeout(()=>document.getElementById('rk-input').focus(),100);
  document.getElementById('rk-input').addEventListener('keydown',e=>{if(e.key==='Enter')executarRanking();});
}

function executarRanking(){
  const tipo=document.getElementById('rk-tipo').value;
  const val=document.getElementById('rk-input').value.trim();
  if(!val){showToast('❌ Informe um termo',true);return;}
  showToast('🏆 Rankeando vendas...');
  const url=tipo==='categoria'?`/api/ranking-categoria?cat=${encodeURIComponent(val)}&limite=30`:`/api/ranking-busca?q=${encodeURIComponent(val)}&limite=30`;
  fetch(url).then(r=>r.json()).then(d=>{
    const res=document.getElementById('rk-result');
    if(!d.ok){res.innerHTML=`<div style="background:#450a0a;color:#f87171;padding:10px;border-radius:5px">❌ ${d.erro}</div>`;return;}
    res.innerHTML=renderRanking(d,tipo);
  }).catch(()=>showToast('❌ Erro',true));
}

function renderRanking(d, tipo){
  const titulo=tipo==='categoria'?`${d.categoria_nome||d.category_id}`:`"${d.query}"`;
  const rows=(d.itens||[]).map((i,idx)=>{
    const medal=idx===0?'🥇':idx===1?'🥈':idx===2?'🥉':'#'+(idx+1);
    return `<tr style="border-bottom:1px solid #1a1f3a">
      <td style="padding:4px 5px;text-align:center;font-weight:700;color:#fbbf24;width:30px">${medal}</td>
      <td style="padding:4px 5px;width:36px">${i.thumbnail?`<img src="${i.thumbnail}" style="width:32px;height:32px;object-fit:cover;border-radius:3px">`:''}</td>
      <td style="padding:4px 7px;max-width:220px">
        <div style="color:#e4e6eb;font-size:.74rem;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${i.titulo}">${i.titulo}</div>
        <div style="color:#8b92a5;font-size:.62rem">${i.seller_apelido||''} ${i.free_shipping?'· 🚚 grátis':''}</div>
      </td>
      <td style="padding:4px 7px;color:#4ade80;text-align:right;font-weight:700">R$ ${fmt(i.preco)}</td>
      <td style="padding:4px 7px;color:#fbbf24;text-align:right;font-weight:800">${i.vendas_total}</td>
      <td style="padding:4px 7px;color:#c084fc;text-align:right;font-weight:700">R$ ${fmt(i.receita_acum)}</td>
      <td style="padding:4px 3px;white-space:nowrap">
        <button onclick="espionarDoRanking('${i.item_id}')" style="background:#78350f;border:none;color:#fbbf24;padding:2px 6px;border-radius:3px;cursor:pointer;font-size:.65rem">🕵️</button>
        <a href="${i.link}" target="_blank" style="color:#667eea;font-size:.7rem;margin-left:2px">↗</a>
      </td>
    </tr>`;
  }).join('');

  return `<div style="background:#0a0e27;border-radius:6px;padding:10px;margin-top:6px">
    <div style="color:#f472b6;font-weight:700;font-size:.88rem;margin-bottom:6px">🏆 Top ${d.retornados} em ${titulo}</div>
    <div style="color:#8b92a5;font-size:.7rem;margin-bottom:8px">${(d.total_resultados_ml||0).toLocaleString('pt-BR')} anúncios no ML · ${d.gerado_em}</div>
    <div style="max-height:420px;overflow-y:auto">
    <table style="width:100%;border-collapse:collapse;font-size:.78rem">
      <thead><tr style="border-bottom:1px solid #2d3452;position:sticky;top:0;background:#0a0e27">
        <th></th><th></th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:left">Anúncio</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Preço</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Vendas</th>
        <th style="padding:5px 7px;color:#8b92a5;text-align:right">Receita est.</th>
        <th></th>
      </tr></thead>
      <tbody>${rows||'<tr><td colspan="7" style="padding:20px;text-align:center;color:#8b92a5">Sem resultados.</td></tr>'}</tbody>
    </table></div>
  </div>`;
}

function espionarDoRanking(itemId){
  abrirEspiao();
  setTimeout(()=>{document.getElementById('spy-input').value=itemId;executarSpy();},200);
}

// Fecha modal com Escape
document.addEventListener('keydown', e => { if(e.key === 'Escape') document.querySelectorAll('.modal-bg').forEach(m=>m.remove()); });
</script>
</body></html>"""


# ════════════════════════════════════════════════════════════════════
# HTML — ADMIN DE USUÁRIOS (tenant)
# ════════════════════════════════════════════════════════════════════
HTML_ADMIN_USUARIOS = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ brand }} — Admin de Usuários</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif}
body{background:#0a0e27;color:#e4e6eb;min-height:100vh}
.topbar{background:#1a1f3a;padding:10px 20px;border-bottom:1px solid #2d3452;
        display:flex;align-items:center;gap:16px}
.logo{font-size:1.4rem;font-weight:900;
     background:linear-gradient(135deg,#667eea,#c084fc);
     -webkit-background-clip:text;-webkit-text-fill-color:transparent}
.spacer{flex:1}
.btn{background:#2d3452;border:none;color:#e4e6eb;padding:7px 14px;
     border-radius:6px;cursor:pointer;font-size:.8rem;text-decoration:none;display:inline-block}
.btn-green{background:linear-gradient(135deg,#10b981,#059669);color:#fff;font-weight:600}
.btn-red{background:#7f1d1d;color:#fca5a5}
.container{max-width:1000px;margin:30px auto;padding:20px}
.card{background:#1a1f3a;border-radius:12px;padding:24px;margin-bottom:20px;border:1px solid #2d3452}
h1{font-size:1.6rem;margin-bottom:6px}
.sub{color:#8b92a5;font-size:.85rem;margin-bottom:20px}
table{width:100%;border-collapse:collapse}
th{text-align:left;padding:10px;color:#8b92a5;font-size:.75rem;text-transform:uppercase;border-bottom:1px solid #2d3452}
td{padding:10px;border-bottom:1px solid #1a1f3a;font-size:.88rem}
input{width:100%;padding:10px;border:1px solid #2d3452;border-radius:6px;
      background:#0a0e27;color:#e4e6eb;font-size:.9rem;margin-bottom:10px}
.grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}
.badge{background:#065f46;color:#4ade80;padding:3px 8px;border-radius:4px;font-size:.7rem}
.badge-user{background:#1e3a5f;color:#60a5fa}
.toast{position:fixed;top:20px;right:20px;padding:12px 18px;border-radius:8px;
       background:#065f46;color:#4ade80;font-weight:600;z-index:9999;opacity:0;transition:opacity .2s}
.toast.show{opacity:1}
.toast.err{background:#450a0a;color:#f87171}
.plano-badge{display:inline-block;padding:3px 8px;border-radius:4px;font-size:.7rem;
             background:linear-gradient(135deg,#667eea,#c084fc);color:#fff;font-weight:700;margin-left:8px}
</style>
</head>
<body>
<div class="topbar">
  <span class="logo" title="{{ plataforma }} · {{ tenant_nome }}">{{ brand }}</span>
  <span style="color:#8b92a5;font-size:.72rem"><span style="color:#c084fc;font-weight:600">{{ tenant_nome }}</span> · {{ usuario }}</span>
  <div class="spacer"></div>
  <a href="/" class="btn">📦 Produtos</a>
  <a href="/ml" class="btn">🚀 Painel ML</a>
  <a href="/logout" class="btn btn-red">Sair</a>
</div>

<div class="container">
  <div class="card">
    <h1>⚙️ Usuários da {{ tenant.nome_empresa }}<span class="plano-badge">{{ tenant.plano }}</span></h1>
    <div class="sub">Convide pessoas do seu time. Dados ficam isolados por empresa.</div>

    <h3 style="font-size:1rem;margin:20px 0 10px;color:#c084fc">➕ Convidar novo usuário</h3>
    <div class="grid">
      <input id="n_email" type="email" placeholder="email@empresa.com">
      <input id="n_nome" type="text" placeholder="Nome">
      <input id="n_senha" type="password" placeholder="Senha temporária">
      <button class="btn btn-green" onclick="convidar()">Adicionar</button>
    </div>

    <h3 style="font-size:1rem;margin:24px 0 10px;color:#c084fc">👥 Usuários ativos</h3>
    <table>
      <thead><tr><th>Email</th><th>Nome</th><th>Role</th><th>Status</th><th></th></tr></thead>
      <tbody id="tb_users"><tr><td colspan="5" style="color:#8b92a5;text-align:center;padding:20px">Carregando...</td></tr></tbody>
    </table>
  </div>

  <div class="card" style="background:#064e3b1a;border-color:#10b981">
    <h3 style="color:#4ade80;font-size:1rem;margin-bottom:8px">🎯 Isolamento multi-tenant</h3>
    <div style="color:#e4e6eb;font-size:.85rem;line-height:1.6">
      • Cada usuário criado aqui só vê os dados da <b>{{ tenant.nome_empresa }}</b><br>
      • O Mercado Livre conectado em "Painel ML" é exclusivo desta empresa<br>
      • Catálogos, espião, watchlist e métricas são 100% isolados de outros clientes do {{ plataforma }}
    </div>
  </div>
</div>

<div id="toast" class="toast"></div>
<script>
function toast(msg, err){
  var t=document.getElementById('toast'); t.textContent=msg;
  t.className='toast show'+(err?' err':''); setTimeout(function(){t.className='toast';},2500);
}
function carregar(){
  fetch('/api/admin/usuarios').then(function(r){return r.json();}).then(function(d){
    if(!d.ok){toast(d.erro||'Erro',true);return;}
    var tb=document.getElementById('tb_users');
    if(!d.usuarios.length){tb.innerHTML='<tr><td colspan="5" style="color:#8b92a5;text-align:center;padding:20px">Sem usuários ainda</td></tr>';return;}
    tb.innerHTML=d.usuarios.map(function(u){
      var roleBadge=u.role==='admin'?'<span class="badge">admin</span>':'<span class="badge badge-user">'+u.role+'</span>';
      var statusBadge=u.ativo?'<span class="badge">ativo</span>':'<span class="badge" style="background:#450a0a;color:#f87171">inativo</span>';
      var btnRem=u.role==='admin'?'':'<button class="btn btn-red" onclick="remover('+u.id+',\\''+u.email+'\\')">Remover</button>';
      return '<tr><td>'+u.email+'</td><td>'+u.nome+'</td><td>'+roleBadge+'</td><td>'+statusBadge+'</td><td>'+btnRem+'</td></tr>';
    }).join('');
  });
}
function convidar(){
  var e=document.getElementById('n_email').value.trim();
  var n=document.getElementById('n_nome').value.trim();
  var s=document.getElementById('n_senha').value;
  if(!e||!s){toast('Email e senha são obrigatórios',true);return;}
  fetch('/api/admin/convidar',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({email:e,nome:n,senha:s,role:'user'})})
  .then(function(r){return r.json();}).then(function(d){
    if(d.ok){toast('✅ Usuário adicionado');['n_email','n_nome','n_senha'].forEach(function(id){document.getElementById(id).value='';});carregar();}
    else toast(d.erro,true);
  });
}
function remover(id,email){
  if(!confirm('Remover acesso de '+email+'?')) return;
  fetch('/api/admin/remover',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({usuario_id:id})})
  .then(function(r){return r.json();}).then(function(d){
    if(d.ok){toast('✅ Usuário removido');carregar();} else toast(d.erro||'Erro',true);
  });
}
carregar();
</script>
</body>
</html>"""


# ════════════════════════════════════════════════════════════════════
# HTML — PÁGINA DE CONFIGURAÇÕES
# ════════════════════════════════════════════════════════════════════
HTML_CONFIG = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ brand }} — Configurações</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0a0e27;color:#e4e6eb;font-size:.85rem}
.topbar{background:#1a1f3a;border-bottom:1px solid #2d3452;padding:10px 20px;display:flex;align-items:center;gap:12px}
.logo{font-size:1.1rem;font-weight:900;background:linear-gradient(135deg,#667eea,#764ba2);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.btn{padding:6px 14px;border:none;border-radius:5px;cursor:pointer;font-size:.78rem;font-weight:600}
.btn-blue{background:#667eea;color:#fff}.btn-green{background:#059669;color:#fff}
.btn-gray{background:#2d3452;color:#e4e6eb}.btn-yellow{background:#f59e0b;color:#000}
.btn-red{background:#dc2626;color:#fff}
.btn:hover{opacity:.85}
.page{max-width:800px;margin:0 auto;padding:24px 20px}
.section{background:#1a1f3a;border-radius:10px;border:1px solid #2d3452;margin-bottom:20px;overflow:hidden}
.section-header{padding:14px 18px;border-bottom:1px solid #2d3452;display:flex;align-items:center;gap:10px}
.section-header h2{font-size:.95rem;font-weight:700}
.section-body{padding:18px}
label{display:block;color:#8b92a5;font-size:.72rem;text-transform:uppercase;margin-bottom:5px}
input,select,textarea{width:100%;background:#0a0e27;border:1px solid #2d3452;color:#e4e6eb;padding:9px 12px;border-radius:6px;font-size:.85rem;margin-bottom:14px}
input:focus,select:focus,textarea:focus{border-color:#667eea;outline:none}
.status-ok{background:#064e3b;color:#4ade80;padding:8px 14px;border-radius:6px;font-size:.82rem;display:flex;align-items:center;gap:8px}
.status-err{background:#450a0a;color:#f87171;padding:8px 14px;border-radius:6px;font-size:.82rem;display:flex;align-items:center;gap:8px}
.upload-zone{background:#0a0e27;border:2px dashed #2d3452;border-radius:8px;padding:32px;text-align:center;cursor:pointer;transition:border-color .2s;margin-bottom:14px}
.upload-zone:hover{border-color:#667eea}
.upload-icon{font-size:2.5rem;margin-bottom:10px}
.progress-bar{height:5px;background:#2d3452;border-radius:3px;margin-top:10px;overflow:hidden;display:none}
.progress-fill{height:100%;background:#667eea;width:0%;transition:width .5s;border-radius:3px}
.info-box{background:#0a0e27;border-radius:6px;padding:10px 14px;font-size:.78rem;color:#8b92a5;margin-bottom:14px}
.row{display:flex;gap:10px;align-items:flex-end}
.row>*{flex:1}
.row>.btn{flex:0 0 auto;margin-bottom:14px}
.result-box{background:#0a0e27;border-radius:6px;padding:12px;font-size:.8rem;margin-top:10px;display:none}
</style>
</head>
<body>
<div class="topbar">
  <span class="logo" title="{{ plataforma }} · {{ tenant_nome }}">{{ brand }}</span>
  <span style="color:#667eea;font-size:.85rem">⚙️ Configurações</span>
  <div style="flex:1"></div>
  <span style="color:#8b92a5;font-size:.72rem"><span style="color:#c084fc;font-weight:600">{{ tenant_nome }}</span> · {{ usuario }}{% if is_admin %} · <a href="/admin/usuarios" style="color:#fbbf24;text-decoration:none">⚙️</a>{% endif %}</span>
  <a href="/" class="btn btn-blue">← Dashboard</a>
  <a href="/logout" class="btn btn-gray">Sair</a>
</div>

<div class="page">

  <!-- ML AUTH -->
  <div class="section">
    <div class="section-header">
      <span style="font-size:1.3rem">🔗</span>
      <div>
        <h2>Conexão Mercado Livre</h2>
        <div style="color:#8b92a5;font-size:.72rem">Token OAuth — renova automaticamente</div>
      </div>
      <div style="margin-left:auto">
        {% if ml_ok %}
        <div class="status-ok">✅ Conectado · User: {{ ml_user }}</div>
        {% else %}
        <div class="status-err">❌ Não conectado</div>
        {% endif %}
      </div>
    </div>
    <div class="section-body">
      {% if ml_ok %}
      <div class="info-box">✅ ML conectado! O token é renovado automaticamente a cada 6h. Você não precisa reconectar.</div>
      <button class="btn btn-gray" onclick="reconectarML()">🔄 Reconectar (novo token)</button>
      {% else %}
      <div class="info-box">Para conectar, clique no link abaixo, autorize o app QUBO no Mercado Livre, e cole o código que aparecer na URL.</div>
      <a href="https://auth.mercadolivre.com.br/authorization?response_type=code&client_id=5055987535998228&redirect_uri=https://www.google.com"
         target="_blank" class="btn btn-yellow" style="display:inline-block;margin-bottom:14px;text-decoration:none">
         1. Clique aqui para autorizar no ML →
      </a>
      <label>2. Cole o código TG-... que apareceu na URL</label>
      <div class="row">
        <input type="text" id="ml-code" placeholder="TG-XXXXXXXXXXX..." style="margin-bottom:0">
        <button class="btn btn-green" onclick="conectarML()" style="margin-bottom:0">Conectar</button>
      </div>
      <div id="ml-result" class="result-box"></div>
      {% endif %}
    </div>
  </div>

  <!-- UPLOAD PDF EM LOTE -->
  <div class="section">
    <div class="section-header">
      <span style="font-size:1.3rem">📤</span>
      <div>
        <h2>Importar Catálogos PDF</h2>
        <div style="color:#8b92a5;font-size:.72rem">Múltiplos PDFs de uma vez — extração com IA (Groq + Mistral)</div>
      </div>
    </div>
    <div class="section-body">
      <div class="upload-zone" id="upload-zone"
           onclick="document.getElementById('pdfInput').click()"
           ondragover="event.preventDefault();this.style.borderColor='#667eea'"
           ondragleave="this.style.borderColor='#2d3452'"
           ondrop="event.preventDefault();handleDrop(event)">
        <div class="upload-icon">📁</div>
        <div style="color:#e4e6eb;font-weight:600;margin-bottom:6px">Clique ou arraste PDFs aqui</div>
        <div style="color:#8b92a5;font-size:.75rem">Selecione um ou vários PDFs · O nome do arquivo vira o fornecedor</div>
      </div>
      <input type="file" id="pdfInput" accept=".pdf" multiple style="display:none" onchange="iniciarFila(this.files)">

      <!-- Barra de progresso geral -->
      <div class="progress-bar" id="progress-bar">
        <div class="progress-fill" id="progress-fill"></div>
      </div>

      <!-- Contador geral -->
      <div id="batch-counter" style="display:none;font-size:.78rem;color:#8b92a5;margin-top:8px;margin-bottom:8px"></div>

      <!-- Fila de arquivos -->
      <div id="fila-upload" style="display:none;margin-top:12px">
        <div style="font-size:.72rem;color:#8b92a5;text-transform:uppercase;margin-bottom:8px;font-weight:600">Fila de processamento</div>
        <div id="fila-itens"></div>
      </div>

      <!-- Resumo final -->
      <div id="upload-result" class="result-box"></div>
    </div>
  </div>

  <!-- VERIFICAR PDFs LOCAL -->
  <div class="section">
    <div class="section-header">
      <span style="font-size:1.3rem">📁</span>
      <div>
        <h2>Pasta de Monitoramento (File Watcher)</h2>
        <div style="color:#8b92a5;font-size:.72rem">Pasta local monitorada pelo watcher automático</div>
      </div>
    </div>
    <div class="section-body">
      <label>Caminho da pasta de PDFs</label>
      <div class="row">
        <input type="text" id="pasta-input" placeholder="C:\\Users\\...\\pasta-dos-catalogos" style="margin-bottom:0"
               value="C:\\Users\\Luiz Gustavo\\OneDrive\\Documents\\Escalada Econ\\Loja QUBO\\Fornecedores\\1 -SISTEMA CATALOGOS AUTOMATICO">
        <button class="btn btn-blue" onclick="salvarPasta()" style="margin-bottom:0">Salvar</button>
      </div>
      <div style="height:10px"></div>
      <button class="btn btn-gray" onclick="verificarPDFs()">🔍 Verificar PDFs na Pasta</button>
      <div id="pasta-result" class="result-box"></div>
    </div>
  </div>

  <!-- INFO -->
  <div class="section">
    <div class="section-header">
      <span style="font-size:1.3rem">ℹ️</span>
      <h2>Informações do Sistema</h2>
    </div>
    <div class="section-body">
      <div class="info-box" style="line-height:1.8">
        <strong>Dashboard:</strong> https://qubo-dashboard.onrender.com<br>
        <strong>Banco:</strong> Supabase PostgreSQL<br>
        <strong>File Watcher:</strong> QUBO_FileWatcher (Tarefa Windows — inicia com o login)<br>
        <strong>Credenciais:</strong> gustavo / qubo2026<br>
        <strong>ML Client ID:</strong> 5055987535998228
      </div>
    </div>
  </div>

</div>

<script>
function showToast(msg, ok=true){
  const box = document.createElement('div');
  box.textContent = msg;
  box.style.cssText = `position:fixed;bottom:20px;right:20px;padding:10px 16px;border-radius:6px;
    font-size:.82rem;font-weight:600;z-index:9999;
    background:${ok?'#064e3b':'#450a0a'};color:${ok?'#4ade80':'#f87171'};
    border:1px solid ${ok?'#4ade80':'#f87171'}`;
  document.body.appendChild(box);
  setTimeout(()=>box.remove(), 3500);
}

function conectarML(){
  const code = document.getElementById('ml-code').value.trim();
  if(!code){ showToast('Cole o código primeiro!', false); return; }
  showToast('🔄 Conectando...');
  fetch('/api/ml-auth',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({code})})
  .then(r=>r.json()).then(d=>{
    const box = document.getElementById('ml-result');
    box.style.display='block';
    if(d.ok){
      box.style.color='#4ade80';
      box.innerHTML='✅ Conectado! User ID: '+d.user_id+'<br>Recarregando...';
      setTimeout(()=>location.reload(), 1500);
    } else {
      box.style.color='#f87171';
      box.innerHTML='❌ '+d.erro;
    }
  }).catch(()=>showToast('❌ Erro de conexão',false));
}

function reconectarML(){
  if(!confirm('Tem certeza? Você precisará autorizar novamente no ML.')) return;
  window.open('https://auth.mercadolivre.com.br/authorization?response_type=code&client_id=5055987535998228&redirect_uri=https://www.google.com','_blank');
}

function handleDrop(e){
  const files = Array.from(e.dataTransfer.files).filter(f=>f.name.toLowerCase().endsWith('.pdf'));
  if(!files.length){ showToast('❌ Apenas arquivos PDF', false); return; }
  iniciarFila(files);
}

function iniciarFila(fileList){
  const files = Array.from(fileList).filter(f=>f.name.toLowerCase().endsWith('.pdf'));
  if(!files.length){ showToast('❌ Nenhum PDF selecionado', false); return; }

  // Reseta UI
  const bar = document.getElementById('progress-bar');
  const fill = document.getElementById('progress-fill');
  const result = document.getElementById('upload-result');
  const counter = document.getElementById('batch-counter');
  const fila = document.getElementById('fila-upload');
  const itens = document.getElementById('fila-itens');

  bar.style.display='block'; fill.style.width='0%';
  result.style.display='none';
  fila.style.display='block';
  itens.innerHTML='';
  counter.style.display='block';
  counter.textContent='Preparando '+files.length+' arquivo(s)...';

  // Cria uma linha por arquivo na fila
  files.forEach((f,i)=>{
    const row = document.createElement('div');
    row.id='fila-item-'+i;
    row.style.cssText='display:flex;align-items:center;gap:10px;padding:7px 10px;border-radius:5px;background:#0a0e27;margin-bottom:6px;font-size:.78rem';
    row.innerHTML=`<span id="fila-icon-${i}" style="font-size:1rem">⏳</span>
      <span style="flex:1;color:#c4c9d4;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${f.name}</span>
      <span id="fila-status-${i}" style="color:#8b92a5;white-space:nowrap">Aguardando...</span>`;
    itens.appendChild(row);
  });

  // Processa sequencialmente
  let totalSalvos=0, ok=0, erros=0;
  (async()=>{
    for(let i=0;i<files.length;i++){
      const f=files[i];
      const icon=document.getElementById('fila-icon-'+i);
      const status=document.getElementById('fila-status-'+i);
      const row=document.getElementById('fila-item-'+i);

      icon.textContent='🔄';
      status.style.color='#667eea';
      status.textContent='Enviando...';
      counter.textContent=`Processando ${i+1} de ${files.length}...`;
      fill.style.width=((i/files.length)*100)+'%';

      try{
        const fd=new FormData(); fd.append('arquivo',f);
        const r=await fetch('/api/upload-pdf',{method:'POST',body:fd});
        const d=await r.json();
        if(d.ok){
          icon.textContent='✅';
          status.style.color='#4ade80';
          status.textContent=d.produtos_extraidos+' produtos ('+d.provider+')';
          row.style.background='#052e16';
          totalSalvos+=d.produtos_extraidos; ok++;
        } else {
          icon.textContent='❌';
          status.style.color='#f87171';
          status.textContent=d.erro||'Erro desconhecido';
          row.style.background='#2d0a0a';
          erros++;
        }
      }catch(e){
        icon.textContent='❌';
        status.style.color='#f87171';
        status.textContent='Erro de conexão';
        row.style.background='#2d0a0a';
        erros++;
      }
    }

    // Resumo final
    fill.style.width='100%';
    counter.style.display='none';
    result.style.display='block';
    if(ok>0 && erros===0){
      result.style.color='#4ade80';
      result.innerHTML=`✅ <strong>${ok} PDF(s)</strong> importados com sucesso · <strong>${totalSalvos} produtos</strong> extraídos!<br><a href="/" style="color:#667eea">Ver no Dashboard →</a>`;
      showToast('✅ '+totalSalvos+' produtos importados!');
    } else if(ok>0){
      result.style.color='#f59e0b';
      result.innerHTML=`⚠️ <strong>${ok} OK</strong> (${totalSalvos} produtos) · <strong>${erros} com erro</strong><br><a href="/" style="color:#667eea">Ver no Dashboard →</a>`;
      showToast('⚠️ '+ok+' OK, '+erros+' com erro', false);
    } else {
      result.style.color='#f87171';
      result.innerHTML='❌ Todos os uploads falharam. Verifique os erros acima.';
      showToast('❌ Falha no upload', false);
    }
    setTimeout(()=>{bar.style.display='none';fill.style.width='0%';},3000);
  })();
}

function salvarPasta(){
  const pasta = document.getElementById('pasta-input').value.trim();
  if(!pasta){ showToast('Digite o caminho da pasta', false); return; }
  fetch('/api/config-pasta',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pasta})})
  .then(r=>r.json()).then(d=>{ if(d.ok) showToast('✅ Pasta salva!'); else showToast('❌ '+d.erro,false); });
}

function verificarPDFs(){
  showToast('🔍 Verificando...');
  fetch('/api/verificar-pdfs').then(r=>r.json()).then(d=>{
    const box = document.getElementById('pasta-result');
    box.style.display='block';
    if(d.ok){
      box.style.color='#4ade80';
      box.innerHTML='📁 <strong>'+d.pasta+'</strong><br>'+
        '📄 PDFs pendentes: <strong>'+d.pdfs+'</strong> | ✅ Enviados: <strong>'+d.enviados+'</strong>'+
        (d.pdfs>0?'<br><br>Arquivos:<br>'+d.arquivos.map(f=>'• '+f).join('<br>'):'');
    } else {
      box.style.color='#f87171';
      box.innerHTML='❌ '+d.erro;
    }
  }).catch(()=>showToast('❌ Erro',false));
}
</script>
</body></html>"""


# ════════════════════════════════════════════════════════════════════
# PÁGINA ESCOLHIDOS (Formação de Preço)
# ════════════════════════════════════════════════════════════════════
@app.route('/escolhidos')
@login_required
def escolhidos():
    tenant = get_tenant_id(); p = ph()
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM produtos WHERE tenant_id = {p} AND escolhido = 1 ORDER BY fornecedor, descricao", [tenant])
    cols = [d[0] for d in cur.description]
    produtos = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    usuario = get_usuario_nome()
    return render_template_string(HTML_ESCOLHIDOS, produtos=produtos, usuario=usuario, total=len(produtos), **_brand_ctx())

@app.route('/exportar-escolhidos')
@login_required
def exportar_escolhidos():
    import openpyxl; from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    tenant = get_tenant_id(); p = ph()
    conn = get_conn(); cur = conn.cursor()
    cur.execute(f"SELECT * FROM produtos WHERE tenant_id = {p} AND escolhido = 1 ORDER BY fornecedor", [tenant])
    cols = [d[0] for d in cur.description]; rows = cur.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Escolhidos QUBO"
    headers = ['Fornecedor','Código','Descrição','Custo','Preço ML','Taxa %','Frete','Custo Total','Margem %','Margem R$','Viável']
    fill = PatternFill(fgColor="1a1f3a", fill_type="solid")
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h); c.fill = fill; c.font = Font(color="FFFFFF", bold=True)
    for row in rows:
        d = dict(zip(cols, row))
        ws.append([d.get('fornecedor'), d.get('codigo'), d.get('descricao'), d.get('custo'),
                   d.get('preco_ml'), round((d.get('taxa_categoria') or 0)*100,1),
                   d.get('custo_frete'), d.get('custo_total'), d.get('margem_percentual'),
                   d.get('margem_reais'), 'SIM' if d.get('viavel') else 'NÃO'])
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'qubo_escolhidos_{datetime.now().strftime("%Y%m%d")}.xlsx')

HTML_ESCOLHIDOS = """<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="UTF-8"><title>{{ brand }} — Escolhidos</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,sans-serif;background:#0a0e27;color:#e4e6eb;font-size:.85rem}
.topbar{background:#1a1f3a;border-bottom:1px solid #2d3452;padding:8px 16px;display:flex;align-items:center;gap:8px;flex-wrap:wrap;position:sticky;top:0}
.logo{font-size:1.1rem;font-weight:900;background:linear-gradient(135deg,#667eea,#764ba2);-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.btn{padding:5px 12px;border:none;border-radius:5px;cursor:pointer;font-size:.78rem;font-weight:600;text-decoration:none;display:inline-block}
.btn-blue{background:#667eea;color:#fff}.btn-green{background:#059669;color:#fff}.btn-gray{background:#2d3452;color:#e4e6eb}
.table-wrap{overflow-x:auto;padding:16px}
table{width:100%;border-collapse:collapse;background:#1a1f3a;border-radius:8px;overflow:hidden;font-size:.78rem}
thead{background:#0f1328}
th,td{padding:7px 8px;border-bottom:1px solid #151929;text-align:left}
th{color:#8b92a5;font-weight:600}
tr:hover td{background:#1f2544}
.badge-ok{background:#064e3b;color:#4ade80;padding:1px 6px;border-radius:8px;font-size:.65rem;font-weight:700}
.badge-no{background:#450a0a;color:#f87171;padding:1px 6px;border-radius:8px;font-size:.65rem;font-weight:700}
</style></head>
<body>
<div class="topbar">
  <span class="logo" title="{{ plataforma }} · {{ tenant_nome }}">{{ brand }}</span>
  <span style="color:#8b92a5">⭐ Escolhidos ({{ total }})</span>
  <div style="flex:1"></div>
  <a href="/exportar-escolhidos" class="btn btn-green">📥 Exportar Excel</a>
  <a href="/" class="btn btn-blue">← Dashboard</a>
</div>
<div class="table-wrap">
<table>
<thead><tr>
  <th>Fornecedor</th><th>Código</th><th>Descrição</th><th>Custo</th><th>Preço ML</th>
  <th>Taxa%</th><th>Frete</th><th>Custo Total</th><th>Margem%</th><th>Margem R$</th><th>Status</th>
</tr></thead>
<tbody>
{% for p in produtos %}
<tr>
  <td style="color:#8b92a5">{{ p.fornecedor }}</td>
  <td style="color:#8b92a5">{{ p.codigo or '' }}</td>
  <td>{{ p.descricao }}</td>
  <td style="color:#fbbf24">R$ {{ p.custo|br }}</td>
  <td style="color:#4ade80;font-weight:700">R$ {{ p.preco_ml|br }}</td>
  <td style="color:#8b92a5">{{ ((p.taxa_categoria or 0.165)*100)|br(1) }}%</td>
  <td>{{ p.custo_frete|br if p.custo_frete else '-' }}</td>
  <td>{{ p.custo_total|br if p.custo_total else '-' }}</td>
  <td style="color:{{ '#4ade80' if (p.margem_percentual or 0) >= 20 else '#f87171' }};font-weight:700">
    {{ p.margem_percentual|br(1) }}%</td>
  <td style="color:{{ '#4ade80' if (p.margem_reais or 0) >= 0 else '#f87171' }}">
    R$ {{ p.margem_reais|br }}</td>
  <td>{% if p.viavel %}<span class="badge-ok">VIÁVEL</span>{% else %}<span class="badge-no">BAIXO</span>{% endif %}</td>
</tr>
{% else %}
<tr><td colspan="11" style="text-align:center;padding:40px;color:#8b92a5">
  Nenhum produto escolhido ainda. <a href="/" style="color:#667eea">Voltar ao Dashboard →</a>
</td></tr>
{% endfor %}
</tbody>
</table>
</div>
</body></html>"""

# ════════════════════════════════════════════════════════════════════
# STARTUP — roda tanto com gunicorn quanto com python direto
# ════════════════════════════════════════════════════════════════════
def _startup():
    """Inicializa banco e config. Erros são logados, nunca crasham o app."""
    try:
        garantir_schema()
        logger.info("Schema OK")
    except Exception as e:
        logger.error(f"garantir_schema falhou: {e}")
    try:
        garantir_config()
        logger.info("Config OK")
    except Exception as e:
        logger.error(f"garantir_config falhou: {e}")

_startup()   # executa ao importar o módulo (gunicorn + python direto)

if __name__ == '__main__':
    PORT = int(os.getenv('PORT', 10000))
    logger.info(f"QUBO Dashboard v4 | 0.0.0.0:{PORT}")
    app.run(host='0.0.0.0', port=PORT, debug=False)
